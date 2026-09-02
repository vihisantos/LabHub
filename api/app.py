import sys, os, re, secrets, hashlib, json, socket, ipaddress, time, functools
from datetime import datetime, timedelta, timezone, date
from io import BytesIO
from urllib.parse import urlparse, parse_qs, quote, urljoin

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'src', 'apps', 'reservalab', 'api')))
from app import app, _SUPABASE_URL, _SUPABASE_SERVICE_KEY, _supabase_headers, _target_subs, push_notify, redis, logger, _is_safe_url, _cache_path
from auth import (
    require_auth,
    require_module as require_module_auth,
    require_cron,
    require_admin,
    require_workspace,
    _verify_jwt,
    _get_token_from_request,
    _get_user_profile,
    _get_workspace,
    _user_in_workspace,
    _is_module_enabled,
    _forbidden,
)
from rbac import (
    require_action as require_action_rbac,
    rbac_enabled as rbac_two_enabled,
    rbac_can as rbac_two_can,
    record_rbac_audit as rbac_record_audit,
)

import requests
from collections import defaultdict
from flask import jsonify, request, g
from openpyxl import load_workbook


# ── Rate limiting (in-memory) ──
_rate_limit_store: dict[str, list[float]] = defaultdict(list)
RATE_LIMIT_WINDOW = 3600  # 1 hora em segundos
RATE_LIMIT_MAX_REQUESTS = 20  # máximo de chamados por IP por hora


def _check_rate_limit(ip: str, max_requests: int | None = None) -> bool:
    """Verifica se o IP excedeu o limite de requisições. Retorna True se permitido.

    ``max_requests`` permite limites específicos por endpoint (padrão:
    RATE_LIMIT_MAX_REQUESTS, usado pelos endpoints existentes).
    """
    limit = RATE_LIMIT_MAX_REQUESTS if max_requests is None else max_requests
    now = datetime.now(timezone.utc).timestamp()
    window_start = now - RATE_LIMIT_WINDOW
    # Remove entradas antigas
    _rate_limit_store[ip] = [t for t in _rate_limit_store[ip] if t > window_start]
    if len(_rate_limit_store[ip]) >= limit:
        return False
    _rate_limit_store[ip].append(now)
    return True


def _require_action_in_handler(action, scope='workspace', resource_type=None, resource_id=None):
    """In-handler RBAC enforcement para rotas cujo workspace só é conhecido após
    resolver o recurso (ex.: Chamados `<id>`).

    O contratante DEVE resolver o workspace real do recurso e setar `g.workspace_id`
    ANTES de chamar (para scope != 'global'). Nunca confiar em workspace fornecido
    pelo cliente. Respeita RBAC_2_ENABLED (OFF ⇒ no-op, legado preservado);
    fail-closed; NUNCA transforma erro em allow. Retorna None se permitido ou
    uma resposta Flask (403) se negado.
    """
    user = getattr(g, 'user', None)
    if not rbac_two_enabled():
        return None
    if not user:
        return _forbidden('Permissão insuficiente')
    scope_normalized = str(scope or 'workspace').strip()
    workspace_id = None if scope_normalized == 'global' else getattr(g, 'workspace_id', None)
    result = rbac_two_can(user, workspace_id, action, scope_normalized)
    rbac_record_audit(
        actor_id=user.get('id'),
        actor_is_super=user.get('is_super_admin'),
        action=action,
        workspace_id=workspace_id,
        scope=scope_normalized,
        effect='allow' if result else 'deny',
        outcome='success' if result else 'denied',
        resource_type=resource_type,
        resource_id=resource_id,
    )
    if not result:
        return _forbidden('Permissão insuficiente')
    return None


# ── Ownership de atendimento (independente do RBAC_2_ENABLED) ────────────────
#
# Estas regras implementam o fluxo operacional definitivo de Chamados e valem
# SEMPRE, com RBAC_2_ENABLED=0 OU =1 (a flag controla a granularidade das Actions
# RBAC 2.0; a regra de ownership é de negócio e nunca é desativada):
#
#   - Técnico comum NÃO pode atribuir/reatribuir responsável (nem escolher outro).
#   - Técnico comum só pode COMEÇAR ATENDIMENTO em chamado SEM responsável.
#   - Depois de atribuído a X, os demais técnicos não podem operar (comentar,
#     mudar status, editar, assumir); apenas o responsável, o líder/assigner e o
#     super admin podem.
#
# "Quem é o líder/assigner" depende do modo:
#   - RBAC ON : quem tem a Action `ticket.assign` no workspace (role/override).
#   - RBAC OFF: profile.role legado == 'admin' OU is_super_admin.
# O "tech comum" é qualquer membro que não seja assigner nem super admin.



def _is_assigner(user, workspace_id):
    """Pode atribuir/reatribuir responsável de chamados no workspace?

    RBAC ON  → Action `ticket.assign` (scope workspace).
    RBAC OFF → profile.role == 'admin' (legado) ou is_super_admin.
    """
    if not user:
        return False
    if user.get('is_super_admin'):
        return True
    if rbac_two_enabled():
        return bool(rbac_two_can(user, workspace_id, 'ticket.assign', 'workspace'))
    return str(user.get('role') or '') == 'admin'


def _ticket_owner(ticket):
    """Id do usuário que detém o atendimento (string vazia ⇒ sem responsável)."""
    return str(ticket.get('assignedToUserId') or '').strip()


def _can_operate_ticket(actor, ticket, workspace_id):
    """Ownership: o ator pode executar operações operacionais no chamado?

    Retorna True quando o chamado não tem responsável, ou o ator é o
    responsável, ou é um assigner (líder/admin/super). O check de assigner só
    é consultado quando o chamado tem responsável de outro — evita custo RBAC
    desnecessário quando ainda não há dono.
    """
    if not actor:
        return False
    if actor.get('is_super_admin'):
        return True
    owner = _ticket_owner(ticket)
    if not owner:
        return True
    if str(actor.get('id') or '') == owner:
        return True
    return _is_assigner(actor, workspace_id)


def _enforce_ownership(actor, ticket, workspace_id, resource_id):
    """Retorna resposta Flask (403) se o ator não pode operar o chamado, senão None."""
    if _can_operate_ticket(actor, ticket, workspace_id):
        return None
    return _forbidden('Este chamado está sendo atendido por outro técnico')


# ── Tracking token público (acesso limitado do professor a um chamado) ──
#
# O professor NÃO tem conta autenticada. Cada chamado gera um token
# criptograficamente aleatório cujo hash SHA-256 fica no banco. O token dá
# acesso SOMENTE ao chamado associado — nunca ao workspace nem a APIs internas.
#
# Rate limiting separado (por IP + hash de token) para os endpoints públicos,
# para mitigar enumeração e abuso. In-memory — adequado a single-worker;
# em multi-worker o Redis seria a evolução natural (já disponível p/ push).

_TRACKING_RATE_STORE: dict[str, list[float]] = defaultdict(list)
TRACKING_RATE_WINDOW = 60             # 60 segundos
TRACKING_RATE_MAX = 30                # 30 requisições por janela por chave (IP/hash)


def _check_tracking_rate_limit(key: str) -> bool:
    """Rate limit específico para endpoints públicos do tracking token."""
    now = datetime.now(timezone.utc).timestamp()
    window_start = now - TRACKING_RATE_WINDOW
    _TRACKING_RATE_STORE[key] = [t for t in _TRACKING_RATE_STORE[key] if t > window_start]
    if len(_TRACKING_RATE_STORE[key]) >= TRACKING_RATE_MAX:
        return False
    _TRACKING_RATE_STORE[key].append(now)
    return True


def require_tracking_token(f):
    """Decorator: valida o tracking token do chamado público.

    O token é lido de:
      - rota ``/api/public/chamados/<tracking_token>`` (view_args)
      - header ``X-Tracking-Token``

    NÃO aceita token por query string (evita vazar segredo em URL/histórico).
    Calcula o hash SHA-256, consulta o chamado pelo hash e expõe SOMENTE
    ``g.tracking_ticket`` (id, workspace_id, status) — escopo mínimo.
    Rejeita com 403 se ausente ou inválido (sem distinguir motivos).
    """
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        token = (
            request.headers.get('X-Tracking-Token')
            or request.view_args.get('tracking_token')
        )
        if not token:
            return _forbidden('Tracking token inválido')

        token_hash = hashlib.sha256(token.encode('utf-8')).hexdigest()

        client_ip = _get_client_ip()
        rate_key = f'{client_ip}:{token_hash}'
        if not _check_tracking_rate_limit(rate_key):
            return jsonify({'error': 'Muitas requisições. Tente novamente mais tarde.'}), 429

        resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/chamados_tickets?tracking_token_hash=eq.{quote(token_hash)}'
            f'&select=id,workspace_id,status',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not resp.ok or not resp.json():
            return _forbidden('Tracking token inválido')

        g.tracking_ticket = resp.json()[0]
        return f(*args, **kwargs)
    return wrapper


def _get_client_ip() -> str:
    """Obtém o IP real do cliente, considerando proxy reverso."""
    forwarded = request.headers.get('X-Forwarded-For', '')
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.remote_addr or 'unknown'


# ── Limites de tamanho de texto ──
MAX_FIELD_LENGTHS = {
    'workspace_id': 50,
    'roomName': 100,
    'reportedBy': 100,
    'reportedByEmail': 150,
    'problemCategory': 100,
    'problemArea': 20,
    'problemDescription': 2000,
    'assignedTo': 100,
    'assignedToUserId': 50,
    'statusNote': 1000,
    'author': 100,
    'content': 2000,
}


def _validate_field_length(key: str, value: str) -> str | None:
    """Valida o comprimento de um campo. Retorna mensagem de erro ou None."""
    max_len = MAX_FIELD_LENGTHS.get(key)
    if max_len and len(value) > max_len:
        return f'Campo {key} muito longo (máximo {max_len} caracteres)'
    return None


# ── Cloudinary (helpers compartilhados) ──

def _cloudinary_public_id(image_url: str) -> str | None:
    """Extrai o public_id de uma URL do Cloudinary (folder/public_id.sem_ext).

    Usa apenas operações de string (sem regex) para evitar ReDoS com URLs
    controladas pelo usuário.
    """
    cloud_name = os.environ.get('VITE_CLOUDINARY_CLOUD_NAME') or os.environ.get('CLOUDINARY_CLOUD_NAME', '')
    if not cloud_name or cloud_name not in image_url:
        return None
    marker = '/image/upload/'
    idx = image_url.find(marker)
    if idx < 0:
        return None
    raw = image_url[idx + len(marker):]
    # Remove o prefixo de versão opcional (ex.: v1234567890/)
    slash = raw.find('/')
    if raw[:1] == 'v' and slash > 0 and raw[1:slash].isdigit():
        raw = raw[slash + 1:]
    # Remove query string e fragmento
    raw = raw.split('?', 1)[0].split('#', 1)[0]
    # Remove a extensão de imagem
    for ext in ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.svg', '.pdf'):
        if raw.lower().endswith(ext):
            raw = raw[:-len(ext)]
            break
    return raw or None


def _cloudinary_destroy(image_url: str) -> bool:
    """Apaga uma imagem do Cloudinary pelo secure_url. Retorna True se o destroy foi aceito."""
    public_id = _cloudinary_public_id(image_url)
    if not public_id:
        return False
    cloud_name = os.environ.get('VITE_CLOUDINARY_CLOUD_NAME') or os.environ.get('CLOUDINARY_CLOUD_NAME', '')
    api_key = os.environ.get('CLOUDINARY_API_KEY', '')
    api_secret = os.environ.get('CLOUDINARY_API_SECRET', '')
    if not cloud_name or not api_key or not api_secret:
        return False
    import base64
    auth = base64.b64encode(f'{api_key}:{api_secret}'.encode()).decode()
    destroy_url = f'https://api.cloudinary.com/v1_1/{cloud_name}/image/destroy'
    try:
        resp = requests.post(
            destroy_url,
            data={'public_id': public_id},
            headers={'Authorization': f'Basic {auth}'},
            timeout=10,
        )
        if not resp.ok:
            return False
        return resp.json().get('result') == 'ok'
    except Exception:
        return False


def _is_valid_photo(value: str) -> bool:
    """Aceita base64 (fallback) ou URL do Cloudinary. Rejeita URLs arbitrárias."""
    value = (value or '').strip()
    if not value:
        return True
    if value.startswith('data:image/'):
        return True
    cloud_name = os.environ.get('VITE_CLOUDINARY_CLOUD_NAME') or os.environ.get('CLOUDINARY_CLOUD_NAME', '')
    if cloud_name and value.startswith(f'https://res.cloudinary.com/{cloud_name}/image/upload/'):
        return True
    return False


# ── TV: YouTube API ──

def iso_duration_to_seconds(duration: str) -> int:
    if not duration or not isinstance(duration, str) or len(duration) > 20:
        return 0
    duration = duration.upper()
    if not duration.startswith('PT'):
        return 0
    total = 0
    remaining = duration[2:]
    for suffix, multiplier in (('H', 3600), ('M', 60), ('S', 1)):
        if suffix in remaining:
            idx = remaining.index(suffix)
            num = remaining[:idx]
            if num.isdigit() and num:
                total += int(num) * multiplier
            remaining = remaining[idx + 1:]
    return total


def parse_youtube_url(url: str) -> dict | None:
    try:
        u = urlparse(url)
        host = u.hostname.replace('www.', '') if u.hostname else ''
        qs = parse_qs(u.query)

        if host in ('youtube.com', 'm.youtube.com'):
            playlist_id = qs.get('list', [None])[0]
            video_id = qs.get('v', [None])[0]
            if playlist_id and video_id:
                return {'type': 'playlist', 'videoId': video_id, 'playlistId': playlist_id}
            elif playlist_id:
                return {'type': 'playlist', 'playlistId': playlist_id}
            elif video_id:
                return {'type': 'video', 'videoId': video_id}
        elif host == 'youtu.be':
            video_id = u.path.lstrip('/').split('/')[0]
            if video_id:
                return {'type': 'video', 'videoId': video_id}
    except Exception:
        pass
    return None


@app.route('/api/tv/youtube/fetch', methods=['POST'])
def tv_youtube_fetch():
    try:
        data = request.get_json()
        url = (data or {}).get('url', '')
        if not url:
            return jsonify({'error': 'URL é obrigatória'}), 400

        parsed = parse_youtube_url(url)
        if not parsed:
            return jsonify({'error': 'URL do YouTube inválida'}), 400

        api_key = os.environ.get('YOUTUBE_API_KEY')
        if not api_key:
            return jsonify({'error': 'YouTube API key não configurada'}), 500

        tracks = []

        if parsed['type'] == 'playlist':
            playlist_id = parsed['playlistId']
            pl_url = (
                f'https://www.googleapis.com/youtube/v3/playlistItems'
                f'?part=snippet,contentDetails&playlistId={playlist_id}'
                f'&maxResults=50&key={api_key}'
            )
            resp = requests.get(pl_url, timeout=15)
            if not resp.ok:
                return jsonify({'error': f'Erro YouTube API: {resp.status_code}'}), 502
            pl_data = resp.json()

            items = pl_data.get('items', [])
            video_ids = [
                item['contentDetails']['videoId']
                for item in items if 'contentDetails' in item
            ]

            if not video_ids:
                return jsonify({'tracks': []})

            vid_url = (
                f'https://www.googleapis.com/youtube/v3/videos'
                f'?part=snippet,contentDetails&id={",".join(video_ids)}&key={api_key}'
            )
            vresp = requests.get(vid_url, timeout=15)
            if vresp.ok:
                vdata = vresp.json()
                vid_map = {}
                for vitem in vdata.get('items', []):
                    vid = vitem.get('id', '')
                    title = vitem.get('snippet', {}).get('title', 'Sem título')
                    dur = iso_duration_to_seconds(
                        vitem.get('contentDetails', {}).get('duration', '')
                    )
                    vid_map[vid] = {'title': title, 'duration': dur}

                for vid in video_ids:
                    info = vid_map.get(vid, {'title': 'Sem título', 'duration': 0})
                    tracks.append({
                        'videoId': vid,
                        'title': info['title'],
                        'duration': info['duration'],
                    })
            else:
                for vid in video_ids:
                    tracks.append({'videoId': vid, 'title': 'Carregando...', 'duration': 0})
        else:
            video_id = parsed['videoId']
            vid_url = (
                f'https://www.googleapis.com/youtube/v3/videos'
                f'?part=snippet,contentDetails&id={video_id}&key={api_key}'
            )
            resp = requests.get(vid_url, timeout=15)
            if not resp.ok:
                return jsonify({'error': f'Erro YouTube API: {resp.status_code}'}), 502
            vdata = resp.json()
            items = vdata.get('items', [])
            if items:
                item = items[0]
                tracks.append({
                    'videoId': video_id,
                    'title': item.get('snippet', {}).get('title', 'Sem título'),
                    'duration': iso_duration_to_seconds(
                        item.get('contentDetails', {}).get('duration', '')
                    ),
                })

        return jsonify({'tracks': tracks})

    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/tv/youtube/search', methods=['POST'])
def tv_youtube_search():
    """Busca músicas no YouTube por nome (sem precisar colar URL)."""
    try:
        data = request.get_json() or {}
        q = str(data.get('q', '')).strip()
        if not q:
            return jsonify({'error': 'Informe o nome da música'}), 400
        max_results = max(1, min(int(data.get('maxResults', 8)), 20))

        api_key = os.environ.get('YOUTUBE_API_KEY')
        if not api_key:
            return jsonify({'error': 'YouTube API key não configurada'}), 500

        url = (
            'https://www.googleapis.com/youtube/v3/search'
            f'?part=snippet&type=video&videoCategoryId=10'
            f'&q={quote(q)}&maxResults={max_results}&key={api_key}'
        )
        resp = requests.get(url, timeout=15)
        if not resp.ok:
            return jsonify({'error': f'Erro YouTube API: {resp.status_code}'}), 502
        data_resp = resp.json()

        results = []
        for item in data_resp.get('items', []):
            snippet = item.get('snippet', {})
            video_id = item.get('id', {}).get('videoId', '')
            if not video_id:
                continue
            thumbnails = snippet.get('thumbnails', {})
            thumb = thumbnails.get('medium') or thumbnails.get('default') or {}
            results.append({
                'videoId': video_id,
                'title': snippet.get('title', 'Sem título'),
                'channel': snippet.get('channelTitle', ''),
                'thumbnail': thumb.get('url', ''),
            })

        return jsonify({'results': results})

    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/tv/calendar/extract', methods=['POST'])
def tv_calendar_extract():
    try:
        data = request.get_json() or {}
        pdf_url = data.get('url', '')
        semester_code = data.get('semester_code', '26/2')
        end_date_str = data.get('end_date', '2026-12-18')

        if not pdf_url or not isinstance(pdf_url, str):
            return jsonify({'error': 'URL do PDF é obrigatória'}), 400

        # Proteção SSRF: bloqueia esquemas não-HTTP, loopback, IPs privados e metadata da cloud
        if not _is_safe_url(pdf_url):
            return jsonify({'error': 'URL inválida ou não permitida (proteção SSRF)'}), 400

        # Download PDF
        resp = requests.get(pdf_url, timeout=30)
        if not resp.ok:
            return jsonify({'error': f'Falha ao baixar PDF: HTTP {resp.status_code}'}), 502

        from pypdf import PdfReader
        import io

        pdf_file = io.BytesIO(resp.content)
        reader = PdfReader(pdf_file)

        full_text = []
        for i, page in enumerate(reader.pages):
            txt = page.extract_text()
            if txt:
                full_text.append(txt)

        text_content = "\n".join(full_text)

        # Parse linear determinístico sem regex para capturar itens no formato "DD - Titulo" ou "DD a DD - Titulo" (elimina ReDoS)
        lines = text_content.split('\n')
        extracted_events = []

        months_map = {
            'janeiro': 1, 'fevereiro': 2, 'março': 3, 'marco': 3, 'abril': 4,
            'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8, 'setembro': 9,
            'outubro': 10, 'novembro': 11, 'dezembro': 12
        }

        current_month = 8 # Padrão 2º semestre (agosto)

        for line in lines:
            line_clean = line.strip()
            if not line_clean:
                continue

            # Verificar se é cabeçalho de mês
            for m_name, m_num in months_map.items():
                if m_name in line_clean.lower():
                    current_month = m_num
                    break

            # Divisão determinística por hífens/travessões sem backtracking polinomial
            day_part = None
            title_part = None
            for sep in (' - ', ' – ', ' — ', '-', '–', '—'):
                if sep in line_clean:
                    left, right = line_clean.split(sep, 1)
                    left = left.strip()
                    right = right.strip()
                    if left and left[:1].isdigit():
                        day_part = left
                        title_part = right
                        break

            if day_part and title_part:
                # Limpar encoding e ruídos comuns
                title_part = title_part.encode('latin1', 'ignore').decode('utf-8', 'ignore') if not any(c in title_part for c in 'ãõçáéíóú') else title_part

                extracted_events.append({
                    'id': f'cal_{len(extracted_events) + 1}',
                    'day_part': day_part,
                    'title': title_part,
                    'month': current_month,
                    'semester_code': semester_code,
                    'is_academic_calendar': True
                })

        expires_at = f"{end_date_str}T23:59:59Z"

        return jsonify({
            'success': True,
            'semester_code': semester_code,
            'expires_at': expires_at,
            'total_events': len(extracted_events),
            'events': extracted_events
        })

    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


# ── TV Corporativa — fonte Excel/SharePoint (fase 1: link anônimo) ────────────
# Credenciais Microsoft NUNCA chegam ao frontend; o download é sempre server-side.
# Evolução para links privados (Graph/app-only) será um PR separado.

TV_SOURCE_MAX_URL_LEN = 2048
TV_SOURCE_MAX_BYTES = 8 * 1024 * 1024  # 8 MB
TV_SOURCE_TIMEOUT = (10, 30)  # (conexão, leitura)
TV_SOURCE_MAX_REDIRECTS = 3
TV_SOURCE_PARSE_ROW_CAP = 5000
TV_SOURCE_PREVIEW_CAP = 100
TV_REFRESH_MIN, TV_REFRESH_MAX = 60, 3600


def _tv_validate_source_url(url):
    """Valida a URL da fonte com proteção SSRF em profundidade.

    Camadas: string/tamanho → HTTPS only → _is_safe_url (herdada do ReservaLab:
    bloqueia localhost, IPs literais privados/loopback/link-local/reservados e
    esquemas não-HTTP) → resolução DNS validando TODOS os IPs retornados.
    """
    if not isinstance(url, str) or not url or len(url) > TV_SOURCE_MAX_URL_LEN:
        return False
    parsed = urlparse(url)
    if parsed.scheme != 'https':
        return False
    if not _is_safe_url(url):
        return False
    host = (parsed.hostname or '').lower()
    try:
        infos = socket.getaddrinfo(host, None)
    except Exception:
        return False
    for info in infos:
        addr = info[4][0]
        try:
            ip = ipaddress.ip_address(addr.split('%')[0])
        except ValueError:
            return False
        if ip.is_private or ip.is_loopback or ip.is_link_local \
                or ip.is_reserved or ip.is_multicast or ip.is_unspecified:
            return False
    return True


def _tv_fetch_source_bytes(url):
    """Baixa o arquivo da fonte com revalidação SSRF a cada hop de redirect.

    Retorna (bytes | None, erro | None). Redirects são seguidos manualmente
    (limite TV_SOURCE_MAX_REDIRECTS) para impedir redirecionamento para hosts
    internos. O download é streamado com teto de tamanho.
    """
    current = url
    for _hop in range(TV_SOURCE_MAX_REDIRECTS + 1):
        if not _tv_validate_source_url(current):
            return None, 'URL inválida ou não permitida (proteção SSRF)'
        try:
            resp = requests.get(
                current, timeout=TV_SOURCE_TIMEOUT, allow_redirects=False, stream=True,
            )
        except requests.exceptions.Timeout:
            return None, 'Tempo esgotado ao contatar a fonte'
        except requests.exceptions.RequestException as exc:
            return None, f'Falha de rede ({exc.__class__.__name__})'
        if resp.status_code in (301, 302, 303, 307, 308):
            loc = resp.headers.get('Location', '')
            resp.close()
            if not loc:
                return None, 'Redirect sem destino'
            current = urljoin(current, loc)
            continue
        if not resp.ok:
            resp.close()
            return None, f'A fonte respondeu HTTP {resp.status_code}'
        chunks, total = [], 0
        try:
            for chunk in resp.iter_content(64 * 1024):
                total += len(chunk)
                if total > TV_SOURCE_MAX_BYTES:
                    return None, 'Arquivo maior que o limite permitido'
                chunks.append(chunk)
        finally:
            resp.close()
        return b''.join(chunks), None
    return None, 'Excesso de redirects'


_TV_FIELD_ALIASES = {
    'title': ('title', 'título', 'titulo', 'evento', 'nome'),
    'date': ('date', 'data', 'início', 'inicio'),
    'endDate': ('enddate', 'end date', 'data final', 'fim'),
    'description': ('description', 'descrição', 'descricao', 'obs', 'observação'),
    'location': ('location', 'local', 'sala'),
    'category': ('category', 'categoria', 'tipo'),
}


def _tv_normalize_header(value):
    return re.sub(r'\s+', ' ', str(value)).strip().lower()


def _tv_parse_date(value):
    """Converte célula de data em ISO yyyy-mm-dd; None quando inválida."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(s[:10], fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _tv_parse_events_xlsx(content, sheet_name=None, field_map=None):
    """Normaliza XLSX em eventos genéricos (sem persistir nada).

    Sobrevive a: aba inexistente/vazia, colunas ausentes, datas inválidas,
    linhas vazias/duplicadas, workbook malformado — linhas ruins viram
    ignoredCount sem derrubar a importação. Duplicatas dentro do arquivo são
    contabilizadas como ignoradas (primeira ocorrência vence).
    """
    result = {'events': [], 'validCount': 0, 'ignoredCount': 0}
    fm = {}
    if isinstance(field_map, dict):
        for key, val in field_map.items():
            if key in _TV_FIELD_ALIASES and isinstance(val, str) and val.strip():
                fm[key] = _tv_normalize_header(val)
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        result['error'] = 'Workbook inválido ou corrompido'
        return result
    names = list(wb.sheetnames)
    if not names:
        wb.close()
        result['error'] = 'Workbook sem abas'
        return result
    if sheet_name:
        if sheet_name not in names:
            wb.close()
            result['error'] = f'Aba "{sheet_name}" não encontrada'
            return result
        ws = wb[sheet_name]
    else:
        ws = wb[names[0]]
    rows = ws.iter_rows(min_row=1, max_row=TV_SOURCE_PARSE_ROW_CAP, values_only=True)
    header = []
    for row in rows:
        header = [_tv_normalize_header(c) if c is not None else '' for c in row]
        break

    def _find_col(field):
        wanted = fm.get(field)
        if wanted and wanted in header:
            return header.index(wanted)
        for alias in _TV_FIELD_ALIASES[field]:
            if alias in header:
                return header.index(alias)
        return -1

    idx = {field: _find_col(field) for field in _TV_FIELD_ALIASES}
    seen = set()
    for row in rows:
        if row is None or all(c is None or str(c).strip() == '' for c in row):
            continue
        title = ''
        if idx['title'] >= 0 and idx['title'] < len(row) and row[idx['title']] is not None:
            title = str(row[idx['title']]).strip()
        date_iso = _tv_parse_date(row[idx['date']]) if idx['date'] >= 0 and idx['date'] < len(row) else None
        if not title or not date_iso:
            result['ignoredCount'] += 1
            continue

        def _cell(field, max_len):
            i = idx[field]
            if i < 0 or i >= len(row) or row[i] is None:
                return ''
            return str(row[i]).strip()[:max_len]

        end_iso = _tv_parse_date(row[idx['endDate']]) if idx['endDate'] >= 0 and idx['endDate'] < len(row) else None
        external_id = hashlib.sha256('|'.join([
            title[:300], date_iso, end_iso or '', _cell('location', 200),
        ]).encode('utf-8')).hexdigest()[:16]
        if external_id in seen:
            result['ignoredCount'] += 1
            continue
        seen.add(external_id)
        event = {
            'externalId': external_id,
            'title': title[:300],
            'date': date_iso,
            'origin': 'sharepoint_excel',
        }
        if end_iso:
            event['endDate'] = end_iso
        description = _cell('description', 500)
        location = _cell('location', 200)
        category = _cell('category', 80)
        if description:
            event['description'] = description
        if location:
            event['location'] = location
        if category:
            event['category'] = category
        result['events'].append(event)
        result['validCount'] += 1
    wb.close()
    return result


def _tv_cache_key(workspace_id, event_source_cfg):
    """Chave inclui workspace E hash da configuração: trocar a URL/aba muda a
    chave, então alterar a configuração invalida o cache naturalmente."""
    cfg_hash = hashlib.sha256(json.dumps(
        event_source_cfg, sort_keys=True, ensure_ascii=False,
    ).encode('utf-8')).hexdigest()[:12]
    return f'tv_source_{workspace_id}_{cfg_hash}'


def _tv_cache_read(cache_key, ttl_seconds):
    """Lê cache Redis/arquivo no padrão ReservaLab. Retorna (fresh, stale)."""
    payload = None
    if redis:
        try:
            raw = redis.get(f'cache:{cache_key}')
            if raw:
                payload = json.loads(raw) if isinstance(raw, str) else raw
        except Exception as exc:
            logger.error("Erro ao ler cache Redis (tv_source): %s", exc)
    if payload is None:
        try:
            path = _cache_path(cache_key)
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as fh:
                    payload = json.load(fh)
        except Exception as exc:
            logger.error("Erro ao ler cache em arquivo (tv_source): %s", exc)
    if not isinstance(payload, dict):
        return None, None
    ts = payload.get('timestamp', 0)
    data = payload.get('data')
    fresh = data if ts and time.time() - ts < ttl_seconds else None
    stale = data if ts and data is not None else None
    return fresh, stale


def _tv_cache_write(cache_key, data, ttl_seconds):
    """Grava cache (Redis + arquivo) com TTL dinâmico limitado por constantes."""
    payload = {'data': data, 'timestamp': time.time()}
    ttl_seconds = max(TV_REFRESH_MIN, min(TV_REFRESH_MAX, int(ttl_seconds)))
    if redis:
        try:
            redis.set(f'cache:{cache_key}', json.dumps(payload, ensure_ascii=False), ex=ttl_seconds)
        except Exception as exc:
            logger.error("Erro ao salvar cache Redis (tv_source): %s", exc)
    try:
        with open(_cache_path(cache_key), 'w', encoding='utf-8') as fh:
            json.dump(payload, fh, ensure_ascii=False)
    except Exception as exc:
        logger.error("Erro ao salvar cache em arquivo (tv_source): %s", exc)


def _get_tv_app_settings(workspace_id):
    """Lê workspace_app_settings do workspace autenticado (service key, server-side).

    O cliente nunca envia settings/workspace como autoridade — o identificador
    vem exclusivamente de g.workspace_id, resolvido pelo require_workspace.
    """
    try:
        url = (
            f"{_SUPABASE_URL}/rest/v1/workspace_app_settings"
            f"?workspace_id=eq.{workspace_id}&app_id=eq.tv&select=settings"
        )
        resp = requests.get(url, headers=_supabase_headers(), timeout=10)
        if resp.ok:
            rows = resp.json() or []
            if rows and isinstance(rows[0].get('settings'), dict):
                return rows[0]['settings']
    except Exception as exc:
        logger.error("Erro ao ler tv app settings: %s", exc)
    return {}


def _tv_source_response(payload, status_code=200):
    body = dict(payload)
    events = body.get('events')
    if isinstance(events, list) and len(events) > TV_SOURCE_PREVIEW_CAP:
        body['totalEvents'] = len(events)
        body['events'] = events[:TV_SOURCE_PREVIEW_CAP]
    return jsonify(body), status_code


@app.route('/api/tv/source/fetch', methods=['POST'])
@require_auth
@require_workspace
@require_module_auth('tv')
def tv_source_fetch():
    """Testa/normaliza a fonte configurada do workspace autenticado.

    Não persiste eventos (sem import neste PR); retorna preview + contagens.
    Em falha de rede serve o último resultado válido (stale) quando houver.
    """
    if not _check_rate_limit(_get_client_ip()):
        return jsonify({'error': 'Muitas requisições. Tente novamente mais tarde.'}), 429
    try:
        # Identificadores do corpo do cliente são ignorados de propósito:
        # autoridade = token JWT + membership (require_workspace) apenas.
        settings = _get_tv_app_settings(g.workspace_id)
        source = settings.get('eventSource') or {}
        if not source.get('enabled') or not source.get('url'):
            return jsonify({'ok': False, 'error': 'Fonte externa não configurada para este workspace'}), 400

        display = settings.get('display') or {}
        try:
            refresh = int(display.get('refreshIntervalSeconds') or 300)
        except (TypeError, ValueError):
            refresh = 300
        ttl = max(TV_REFRESH_MIN, min(TV_REFRESH_MAX, refresh))
        cache_key = _tv_cache_key(g.workspace_id, source)

        url = source['url']
        if not _tv_validate_source_url(url):
            return jsonify({'ok': False, 'error': 'URL configurada é inválida ou não permitida (proteção SSRF)'}), 400

        # Dentro do TTL responde do cache sem tocar a rede.
        fresh, _stale = _tv_cache_read(cache_key, ttl)
        if fresh:
            return _tv_source_response({'ok': True, 'freshness': 'fresh', **fresh})

        content, fetch_error = _tv_fetch_source_bytes(url)
        synced_at = datetime.now(timezone.utc).isoformat()

        if fetch_error is None and content:
            parsed = _tv_parse_events_xlsx(content, source.get('sheetName'), source.get('fieldMap'))
            if parsed.get('error'):
                _fresh, stale = _tv_cache_read(cache_key, ttl)
                if stale:
                    return _tv_source_response({
                        'ok': True, 'freshness': 'stale', 'warning': parsed['error'], **stale,
                    })
                return jsonify({'ok': False, 'error': parsed['error']}), 502
            payload = {
                'events': parsed['events'],
                'validCount': parsed['validCount'],
                'ignoredCount': parsed['ignoredCount'],
                'syncedAt': synced_at,
                'source': 'sharepoint_excel',
            }
            _tv_cache_write(cache_key, payload, ttl)
            return _tv_source_response({'ok': True, 'freshness': 'fresh', **payload})

        _fresh, stale = _tv_cache_read(cache_key, ttl)
        if stale:
            return _tv_source_response({
                'ok': True, 'freshness': 'stale', 'warning': fetch_error, **stale,
            })
        return jsonify({'ok': False, 'error': fetch_error or 'Falha ao obter a fonte'}), 502

    except Exception as exc:
        logger.error("Erro em /api/tv/source/fetch: %s", exc)
        return jsonify({'ok': False, 'error': 'Erro interno'}), 500


@app.route('/api/tv/youtube/live', methods=['GET'])
def tv_youtube_live():
    """
    Verifica se o canal da faculdade está ao vivo no YouTube.
    Usa um channel_id fixo ou configurado como variável de ambiente.
    """
    try:
        api_key = os.environ.get('YOUTUBE_API_KEY')
        if not api_key:
            return jsonify({'isLive': False, 'error': 'API key não configurada'}), 200

        channel_id = os.environ.get('YOUTUBE_CHANNEL_ID', '')
        if not channel_id:
            return jsonify({'isLive': False, 'error': 'Channel ID não configurado'}), 200

        # YouTube Data API v3: search for active live broadcasts
        search_url = (
            f'https://www.googleapis.com/youtube/v3/search'
            f'?part=snippet&channelId={channel_id}'
            f'&eventType=live&type=video'
            f'&order=date&maxResults=1&key={api_key}'
        )

        resp = requests.get(search_url, timeout=10)
        if not resp.ok:
            return jsonify({'isLive': False, 'error': f'YouTube API erro: {resp.status_code}'}), 200

        data = resp.json()
        items = data.get('items', [])

        if not items:
            return jsonify({'isLive': False})

        live_item = items[0]
        snippet = live_item.get('snippet', {})
        video_id = live_item.get('id', {}).get('videoId', '')

        # Buscar estatísticas (viewer count)
        stats_url = (
            f'https://www.googleapis.com/youtube/v3/videos'
            f'?part=liveStreamingDetails,snippet'
            f'&id={video_id}&key={api_key}'
        )
        stats_resp = requests.get(stats_url, timeout=10)
        viewer_count = None
        if stats_resp.ok:
            stats_data = stats_resp.json()
            stats_items = stats_data.get('items', [])
            if stats_items:
                live_details = stats_items[0].get('liveStreamingDetails', {})
                viewer_count = live_details.get('concurrentViewers')
                if viewer_count is not None:
                    viewer_count = int(viewer_count)

        return jsonify({
            'isLive': True,
            'channelTitle': snippet.get('channelTitle', ''),
            'videoId': video_id,
            'title': snippet.get('title', ''),
            'thumbnailUrl': snippet.get('thumbnails', {}).get('high', {}).get('url', ''),
            'viewerCount': viewer_count,
        })

    except Exception as e:
        logger.error("Erro em tv_youtube_live: %s", e)
        return jsonify({'isLive': False, 'error': 'Erro interno'}), 200


@app.route('/api/tv/cloudinary/delete', methods=['POST'])
@require_auth
@require_workspace
@require_module_auth('tv')
@require_action_rbac('tv.content.manage', scope='workspace')
def tv_cloudinary_delete():
    """
    Deleta uma imagem do Cloudinary pelo seu secure_url.
    Requer CLOUDINARY_API_KEY e CLOUDINARY_API_SECRET configurados no ambiente.
    Auth required, TV module required.
    """
    try:
        data = request.get_json() or {}
        image_url = data.get('image_url', '')
        if not image_url:
            return jsonify({'success': False, 'error': 'image_url é obrigatório'}), 400

        cloud_name = os.environ.get('VITE_CLOUDINARY_CLOUD_NAME') or os.environ.get('CLOUDINARY_CLOUD_NAME', '')

        if not cloud_name:
            return jsonify({'success': False, 'error': 'Cloudinary cloud_name não configurado'}), 200

        if not os.environ.get('CLOUDINARY_API_KEY') or not os.environ.get('CLOUDINARY_API_SECRET'):
            return jsonify({'success': False, 'error': 'Cloudinary API key/secret não configurados'}), 200

        public_id = _cloudinary_public_id(image_url)
        if not public_id:
            return jsonify({'success': False, 'error': 'URL não é uma imagem do Cloudinary válida'}), 200

        destroyed = _cloudinary_destroy(image_url)
        return jsonify({
            'success': destroyed,
            'result': 'ok' if destroyed else 'not_found',
            'public_id': public_id,
        })

    except Exception as e:
        logger.error("Erro em tv_cloudinary_delete: %s", e)
        return jsonify({'success': False, 'error': 'Erro interno'}), 200


@app.route('/api/tv/health', methods=['GET'])
def tv_health():
    api_key_configured = bool(os.environ.get('YOUTUBE_API_KEY'))
    channel_configured = bool(os.environ.get('YOUTUBE_CHANNEL_ID'))
    return jsonify({
        'status': 'ok',
        'youtube_api_key_configured': api_key_configured,
        'youtube_channel_configured': channel_configured,
    })


# ── TV: Código de ativação do app desktop ──

def _generate_activation_code():
    """Código de 6 caracteres sem caracteres ambíguos (0/O, 1/I)."""
    alphabet = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'
    return ''.join(secrets.choice(alphabet) for _ in range(6))


def _require_supabase():
    if not _SUPABASE_URL or not _SUPABASE_SERVICE_KEY:
        return None
    return True


# ── TV: identidade do kiosk (usuário GoTrue sem senha + sessão via token_hash) ──

_TV_DEVICE_EMAIL_DOMAIN = 'devices.labhub.local'
_uuid_re = re.compile(r'^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$')


def _tv_device_email(device_id: str) -> str:
    """E-mail determinístico do usuário GoTrue do kiosk (sem senha permanente)."""
    return f'kiosk-{device_id.lower()}@{_TV_DEVICE_EMAIL_DOMAIN}'


def _provision_tv_device_session(device_id: str):
    """
    Cria (ou reutiliza) o usuário GoTrue do kiosk e gera um token_hash de
    magiclink de uso único. O app desktop troca o token por sessão com
    supabase.auth.verifyOtp({ token_hash, type: 'magiclink' }).

    Retorna (token_hash, auth_user_id). Levanta RuntimeError em falha.
    """
    email = _tv_device_email(device_id)

    # 1) Usuário SEM senha (não autentica por password; só via magiclink).
    #    422 "already registered" é aceitável: reativação do mesmo device.
    create_resp = requests.post(
        f'{_SUPABASE_URL}/auth/v1/admin/users',
        headers={**_supabase_headers(), 'Content-Type': 'application/json'},
        json={
            'email': email,
            'email_confirm': True,
            'user_metadata': {'role': 'tv_device', 'device_id': device_id},
        },
        timeout=10,
    )
    if create_resp.status_code not in (200, 201, 422):
        raise RuntimeError(f'Falha ao criar identidade da TV ({create_resp.status_code})')

    # 2) Magic link server-side → extrai token_hash do action_link.
    link_resp = requests.post(
        f'{_SUPABASE_URL}/auth/v1/admin/generate_link',
        headers={**_supabase_headers(), 'Content-Type': 'application/json'},
        json={'type': 'magiclink', 'email': email},
        timeout=10,
    )
    if not link_resp.ok:
        raise RuntimeError(f'Falha ao gerar sessão da TV ({link_resp.status_code})')
    payload = link_resp.json() or {}
    action_link = (payload.get('properties') or {}).get('action_link') or ''
    token_hash = (parse_qs(urlparse(action_link).query).get('token') or [None])[0]
    if not token_hash:
        raise RuntimeError('action_link sem token')
    auth_user_id = (payload.get('user') or {}).get('id')
    if not auth_user_id:
        raise RuntimeError('generate_link sem user.id')
    return token_hash, auth_user_id


def _upsert_tv_device_row(device_id: str, name: str, workspace_id: str, auth_user_id: str) -> bool:
    """Registra/atualiza a linha do device (service_role; RLS não se aplica)."""
    up_resp = requests.post(
        f'{_SUPABASE_URL}/rest/v1/tv_devices',
        headers={**_supabase_headers(), 'Prefer': 'resolution=merge-duplicates,return=minimal'},
        json={
            'id': device_id,
            'name': name,
            'workspace_id': workspace_id,
            'user_id': auth_user_id,
            'last_seen': datetime.now(timezone.utc).isoformat(),
        },
        timeout=10,
    )
    return up_resp.ok


def _validate_device_id(device_id):
    """Valida UUID vindo do cliente; retorna string normalizada ou None."""
    if not isinstance(device_id, str) or not _uuid_re.match(device_id.strip()):
        return None
    return device_id.strip().lower()


def require_module(workspace, module_id):
    """Verifica se module_id está habilitado no workspace.

    Retorna None se permitido, ou uma tupla (flask.Response, int) se bloqueado.
    ``workspace`` deve ser o dict já carregado do Supabase (contendo
    ``disabled_apps`` ou não).  Fail-closed: se workspace for None ou vazio,
    o módulo não pode ser verificaçado ⇒ bloqueia (403). A autorização
    nunca deve virar fail-open por falta de contexto de workspace.
    """
    if not workspace:
        return jsonify({
            'error': 'MODULE_NOT_VERIFIABLE',
            'module': module_id,
            'message': 'Não foi possível verificar o módulo para este workspace.',
        }), 403
    disabled = workspace.get('disabled_apps') or []
    if module_id in disabled:
        return jsonify({
            'error': 'MODULE_DISABLED',
            'module': module_id,
            'message': 'Este módulo não está habilitado neste workspace.',
        }), 403
    return None


@app.route('/api/tv/activation/create', methods=['POST'])
def tv_activation_create():
    """
    Gera um código de ativação para o app desktop da TV.
    Requer o token de acesso do usuário logado (Supabase) via Bearer.
    O código é vinculado ao primeiro workspace do usuário (ou ao escolhido,
    se super admin), tem validade de 24h e uso único.
    """
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        token = (request.headers.get('Authorization') or '').replace('Bearer ', '').strip()
        if not token:
            return jsonify({'error': 'Token de autenticação ausente'}), 401

        # 1. Valida o JWT do usuário via Supabase Auth
        auth_resp = requests.get(
            f'{_SUPABASE_URL}/auth/v1/user',
            headers={'apikey': _SUPABASE_SERVICE_KEY, 'Authorization': f'Bearer {token}'},
            timeout=10,
        )
        if not auth_resp.ok:
            return jsonify({'error': 'Sessão inválida ou expirada. Faça login novamente.'}), 401
        auth_user = auth_resp.json()
        user_id = auth_user.get('id')
        if not user_id:
            return jsonify({'error': 'Usuário não identificado'}), 401

        # 2. Perfil do usuário (workspaces atribuídos)
        prof_resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/profiles?id=eq.{quote(user_id)}',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not prof_resp.ok or not prof_resp.json():
            return jsonify({'error': 'Perfil não encontrado'}), 404
        profile = prof_resp.json()[0]
        workspace_ids = profile.get('workspace_ids') or []
        is_super_admin = bool(profile.get('is_super_admin'))

        # 3. Workspace alvo
        body = request.get_json() or {}
        workspace_id = None
        if is_super_admin and body.get('workspace_id'):
            workspace_id = body.get('workspace_id')
        elif workspace_ids:
            workspace_id = workspace_ids[0]
        if not workspace_id:
            return jsonify({'error': 'Este usuário não tem workspace atribuído'}), 400

        ws_resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/workspaces?id=eq.{quote(workspace_id)}',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not ws_resp.ok or not ws_resp.json():
            return jsonify({'error': 'Workspace não encontrado'}), 400

        device_name = str(body.get('device_name') or '').strip()[:60] or None

        # 4. Gera código único (tentativas em caso de colisão)
        expires_at = (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat()
        inserted = None
        for _ in range(5):
            code = _generate_activation_code()
            ins_resp = requests.post(
                f'{_SUPABASE_URL}/rest/v1/tv_activation_codes',
                headers={**_supabase_headers(), 'Prefer': 'return=representation'},
                json={
                    'code': code,
                    'workspace_id': workspace_id,
                    'user_id': user_id,
                    'device_name': device_name,
                    'status': 'pending',
                    'expires_at': expires_at,
                },
                timeout=10,
            )
            if ins_resp.ok:
                inserted = ins_resp.json()[0]
                break
            # PostgREST devolve 409 em violação de unique (colisão de código)
            if ins_resp.status_code != 409:
                return jsonify({'error': f'Erro ao criar o código: {ins_resp.status_code}'}), 500
        if not inserted:
            return jsonify({'error': 'Não foi possível gerar um código único'}), 500

        return jsonify({
            'success': True,
            'code': inserted.get('code'),
            'expires_at': inserted.get('expires_at'),
            'workspace_id': workspace_id,
            'device_name': device_name,
        })

    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/tv/activation/redeem', methods=['POST'])
def tv_activation_redeem():
    """
    Valida e consome um código de ativação (chamado pelo app desktop, anon).
    Provisiona a identidade do kiosk (usuário GoTrue sem senha + vínculo em
    tv_devices) e retorna token_hash para o device obter sessão própria.
    """
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        if not _check_rate_limit(f'tv-redeem:{_get_client_ip()}'):
            return jsonify({'error': 'Muitas tentativas. Aguarde alguns minutos.'}), 429

        body = request.get_json() or {}
        code = str(body.get('code') or '').strip().upper()
        if not code:
            return jsonify({'error': 'Informe o código de ativação'}), 400
        device_id = _validate_device_id(body.get('device_id'))
        if not device_id:
            return jsonify({'error': 'device_id inválido'}), 400

        resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/tv_activation_codes?code=eq.{quote(code)}&select=*',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not resp.ok:
            return jsonify({'error': 'Erro ao validar o código'}), 502
        rows = resp.json()
        if not rows:
            return jsonify({'error': 'Código inválido. Verifique e tente novamente.'}), 404

        row = rows[0]
        if row.get('status') != 'pending':
            return jsonify({'error': 'Código já utilizado'}), 400

        expires_at = row.get('expires_at')
        if expires_at:
            try:
                exp = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
                if exp < datetime.now(timezone.utc):
                    return jsonify({'error': 'Código expirado. Gere um novo no painel.'}), 400
            except Exception:
                pass

        workspace_id = row.get('workspace_id')
        ws_resp = requests.get(
            f"{_SUPABASE_URL}/rest/v1/workspaces?id=eq.{quote(workspace_id)}",
            headers=_supabase_headers(),
            timeout=10,
        )
        workspace = ws_resp.json()[0] if ws_resp.ok and ws_resp.json() else None
        if not workspace:
            return jsonify({'error': 'Workspace do código não encontrado'}), 500

        # Provisiona identidade ANTES de consumir o código: se a infra falhar,
        # o código permanece utilizável.
        try:
            token_hash, auth_user_id = _provision_tv_device_session(device_id)
        except RuntimeError as e:
            logger.error("Erro ao provisionar identidade da TV: %s", e)
            return jsonify({'error': str(e)}), 502

        device_name = (
            str(body.get('device_name') or '').strip()[:60]
            or (row.get('device_name') or '').strip()
            or 'TV Desktop'
        )
        if not _upsert_tv_device_row(device_id, device_name, workspace_id, auth_user_id):
            return jsonify({'error': 'Falha ao registrar o dispositivo'}), 502

        # Consome o código (uso único)
        requests.patch(
            f"{_SUPABASE_URL}/rest/v1/tv_activation_codes?id=eq.{quote(row['id'])}",
            headers={**_supabase_headers(), 'Prefer': 'return=minimal'},
            json={'status': 'used', 'used_at': datetime.now(timezone.utc).isoformat()},
            timeout=10,
        )

        return jsonify({
            'success': True,
            'code': code,
            'workspace': workspace,
            'device_name': device_name,
            'token_hash': token_hash,
        })

    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/tv/devices/provision', methods=['POST'])
def tv_device_provision():
    """
    Provisiona identidade + sessão de kiosk a partir do painel web
    (fluxo de configuração com login humano). Requer Bearer do usuário;
    autorização igual à geração de códigos (super admin em qualquer
    workspace; membro apenas no próprio).
    """
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        token = _get_token_from_request()
        if not token:
            return jsonify({'error': 'Token de autenticação ausente'}), 401

        auth_resp = requests.get(
            f'{_SUPABASE_URL}/auth/v1/user',
            headers={'apikey': _SUPABASE_SERVICE_KEY, 'Authorization': f'Bearer {token}'},
            timeout=10,
        )
        if not auth_resp.ok:
            return jsonify({'error': 'Sessão inválida ou expirada. Faça login novamente.'}), 401
        user_id = (auth_resp.json() or {}).get('id')
        if not user_id:
            return jsonify({'error': 'Usuário não identificado'}), 401

        prof_resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/profiles?id=eq.{quote(user_id)}',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not prof_resp.ok or not prof_resp.json():
            return jsonify({'error': 'Perfil não encontrado'}), 404
        profile = prof_resp.json()[0]
        workspace_ids = profile.get('workspace_ids') or []
        is_super_admin = bool(profile.get('is_super_admin'))

        body = request.get_json() or {}
        workspace_id = body.get('workspace_id')
        device_id = _validate_device_id(body.get('device_id'))
        if not workspace_id or not device_id:
            return jsonify({'error': 'workspace_id e device_id são obrigatórios'}), 400

        if not is_super_admin and workspace_id not in workspace_ids:
            return jsonify({'error': 'Sem permissão neste workspace'}), 403

        ws_resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/workspaces?id=eq.{quote(workspace_id)}&select=id,name,slug,location',
            headers=_supabase_headers(),
            timeout=10,
        )
        workspace = ws_resp.json()[0] if ws_resp.ok and ws_resp.json() else None
        if not workspace:
            return jsonify({'error': 'Workspace não encontrado'}), 400

        try:
            token_hash, auth_user_id = _provision_tv_device_session(device_id)
        except RuntimeError as e:
            logger.error("Erro ao provisionar identidade da TV: %s", e)
            return jsonify({'error': str(e)}), 502

        device_name = str(body.get('device_name') or '').strip()[:60] or 'TV Desktop'
        if not _upsert_tv_device_row(device_id, device_name, workspace_id, auth_user_id):
            return jsonify({'error': 'Falha ao registrar o dispositivo'}), 502

        return jsonify({
            'success': True,
            'workspace': workspace,
            'device_name': device_name,
            'token_hash': token_hash,
        })

    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


# ── TV: snapshot TV-safe do Dashboard de Chamados ──
#
# GET /api/tv/chamados/display — device-only. Fornece à futura CallsDashboardScreen
# um snapshot pequeno e sem PII dos chamados do workspace do dispositivo.
#
# Autorização: JWT → user_id → tv_devices.user_id → workspace_id (vínculo
# persistido no servidor). Nenhum parâmetro do cliente influencia o escopo.
# Projeção: allowlist explícita de colunas na consulta E no serializer — o
# objeto bruto de chamados_tickets NUNCA sai do backend.
# Superfície pública de TV (qualquer pessoa pode ver/fotografar a tela):
# "se o dado não precisa aparecer na TV, não sai do backend".
#
# NÃO é acesso administrativo a chamados: sem escrita, sem comentários,
# sem dados de contato, sem descrição livre, sem patrimônio, sem fotos.

TV_CHAMADOS_TICKET_LIMIT = 100        # máx. de chamados ativos na lista da TV
TV_CHAMADOS_METRICS_WINDOW_DAYS = 30  # janela explícita do histórico p/ métricas
TV_CHAMADOS_RATE_LIMIT_PER_HOUR = 240  # polling legítimo: 1 req/30s = 120/h + margem

_CHAMADOS_ACTIVE_STATUSES = ('aberto', 'a_caminho', 'em_atendimento')
_CHAMADOS_HIGH_PRIORITIES = ('alta', 'urgente')

# Allowlist de colunas — nunca `select=*` nesta superfície.
_TV_CHAMADOS_TICKET_COLS = (
    'ticketNumber,roomName,problemArea,problemCategory,priority,status,createdAt,resolvedAt'
)
_TV_CHAMADOS_METRIC_COLS = 'status,priority,createdAt,resolvedAt,feedbackRating'


def _resolve_tv_device_workspace(user_id) -> str | None:
    """Resolve o workspace do kiosk pelo vínculo persistido tv_devices.user_id.

    Retorna o workspace_id ou None quando a sessão não corresponde a um
    dispositivo registrado ou o device está sem vínculo válido de workspace.
    """
    try:
        resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/tv_devices'
            f'?user_id=eq.{quote(str(user_id))}&select=id,workspace_id&limit=1',
            headers=_supabase_headers(),
            timeout=10,
        )
        rows = resp.json() if resp.ok else []
    except Exception:
        return None
    if not rows:
        return None
    workspace_id = rows[0].get('workspace_id')
    return str(workspace_id) if workspace_id else None


def _tv_project_ticket(row: dict) -> dict:
    """Serializer allowlist do chamado para TV — campos explícitos, nunca bruto.

    Proibidos por design (presentes no banco, ausentes aqui): reportedBy,
    reportedByEmail, problemDescription, assetPatrimony/asset*, photos,
    statusNote, assignedTo*, feedbackComment/feedbackAt/feedbackRating,
    id, roomId, workspace_id e metadados.
    """
    return {
        'ticketNumber': row.get('ticketNumber'),
        'roomName': row.get('roomName') or '',
        'problemArea': row.get('problemArea') or '',
        'problemCategory': row.get('problemCategory') or '',
        'priority': row.get('priority') or 'normal',
        'status': row.get('status') or 'aberto',
        'createdAt': row.get('createdAt'),
        'resolvedAt': row.get('resolvedAt'),
    }


def _tv_chamados_summary(active_rows: list, window_rows: list) -> dict:
    """Métricas agregadas sem nomes: contagens da fila ativa + janela de histórico."""
    open_count = sum(1 for t in active_rows if t.get('status') == 'aberto')
    in_progress = sum(1 for t in active_rows if t.get('status') in _CHAMADOS_ACTIVE_STATUSES[1:])
    high_priority = sum(1 for t in active_rows if t.get('priority') in _CHAMADOS_HIGH_PRIORITIES)

    resolution_total_s = 0.0
    resolution_count = 0
    rating_sum = 0.0
    rating_count = 0
    for t in window_rows:
        if t.get('resolvedAt') and t.get('createdAt'):
            created = _parse_iso(t['createdAt'], 'createdAt')
            resolved = _parse_iso(t['resolvedAt'], 'resolvedAt')
            if created and resolved:
                resolution_total_s += max(0.0, (resolved - created).total_seconds())
                resolution_count += 1
        rating = t.get('feedbackRating')
        if rating:
            try:
                rating_sum += float(rating)
                rating_count += 1
            except (TypeError, ValueError):
                pass

    return {
        'total': len(active_rows),
        'open': open_count,
        'inProgress': in_progress,
        'highPriority': high_priority,
        'avgResolutionHours': round(resolution_total_s / resolution_count / 3600, 1) if resolution_count else None,
        'satisfaction': round(rating_sum / rating_count, 2) if rating_count else None,
    }


@app.route('/api/tv/chamados/display', methods=['GET'])
@require_auth
def tv_chamados_display():
    """Snapshot TV-safe dos chamados do workspace do dispositivo autenticado.

    Device-only (403 para qualquer sessão sem linha em tv_devices com
    workspace válido — inclusive humanos/admins). Consultas sempre escopadas
    pelo vínculo server-side; parâmetros do cliente são ignorados.
    Resposta: { generatedAt, summary, tickets } — sem workspace_id.
    """
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    if not _check_rate_limit(f'tv-chamados-display:{_get_client_ip()}', TV_CHAMADOS_RATE_LIMIT_PER_HOUR):
        return jsonify({'error': 'Muitas requisições. Aguarde alguns instantes.'}), 429
    try:
        workspace_id = _resolve_tv_device_workspace(g.user_id)
        if not workspace_id:
            return jsonify({'error': 'Sessão não corresponde a um dispositivo TV válido'}), 403

        now = datetime.now(timezone.utc)
        cutoff = (now - timedelta(days=TV_CHAMADOS_METRICS_WINDOW_DAYS)).isoformat()
        base = f'{_SUPABASE_URL}/rest/v1/chamados_tickets'
        ws_filter = f'workspace_id=eq.{quote(workspace_id)}&archived=eq.false'

        # 1) Fila ativa (chamados que aparecem na TV)
        actives_resp = requests.get(
            f'{base}?select={_TV_CHAMADOS_TICKET_COLS}&{ws_filter}'
            f'&status=in.({",".join(_CHAMADOS_ACTIVE_STATUSES)})'
            f'&order=createdAt.desc&limit={TV_CHAMADOS_TICKET_LIMIT}',
            headers=_supabase_headers(),
            timeout=10,
        )
        # 2) Janela recente p/ métricas agregadas (saída numérica, sem nomes)
        window_resp = requests.get(
            f'{base}?select={_TV_CHAMADOS_METRIC_COLS}&{ws_filter}'
            f'&createdAt=gte.{quote(cutoff)}&order=createdAt.desc&limit=1000',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not actives_resp.ok or not window_resp.ok:
            return jsonify({'error': 'Erro ao consultar chamados'}), 502

        active_rows = actives_resp.json() or []
        window_rows = window_resp.json() or []
        return jsonify({
            'generatedAt': now.isoformat(),
            'summary': _tv_chamados_summary(active_rows, window_rows),
            'tickets': [_tv_project_ticket(row) for row in active_rows],
        })

    except Exception as exc:
        logger.error("Erro em /api/tv/chamados/display: %s", exc)
        return jsonify({'error': 'Erro interno'}), 500


# ── Chamados (formulário público via QR) ──

CHAMADOS_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS public.chamados_tickets (
    id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    "workspace_id" UUID REFERENCES public.workspaces(id) ON DELETE CASCADE,
    "roomId" TEXT NOT NULL DEFAULT '',
    "roomName" TEXT NOT NULL DEFAULT '',
    "assetId" TEXT DEFAULT '',
    "assetSource" TEXT DEFAULT '',
    "assetName" TEXT DEFAULT '',
    "assetPatrimony" TEXT DEFAULT '',
    "problemCategory" TEXT NOT NULL DEFAULT '',
    "problemArea" TEXT NOT NULL DEFAULT '',
    "problemDescription" TEXT NOT NULL DEFAULT '',
    status TEXT NOT NULL DEFAULT 'aberto',
    "priority" TEXT NOT NULL DEFAULT 'normal',
    "reportedBy" TEXT NOT NULL DEFAULT '',
    "reportedByEmail" TEXT DEFAULT '',
    "assignedTo" TEXT DEFAULT '',
    "assignedToUserId" TEXT DEFAULT '',
    "ticketNumber" INTEGER NOT NULL DEFAULT 0,
    "createdAt" TIMESTAMPTZ DEFAULT NOW(),
    "updatedAt" TIMESTAMPTZ DEFAULT NOW(),
    "resolvedAt" TIMESTAMPTZ,
    "archived" BOOLEAN NOT NULL DEFAULT FALSE,
    "closedAt" TIMESTAMPTZ,
    "closedBy" TEXT DEFAULT '',
    "statusNote" TEXT DEFAULT '',
    "photos" TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS public.ticket_events (
    id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    "ticket_id" UUID NOT NULL REFERENCES public.chamados_tickets(id) ON DELETE CASCADE,
    "workspace_id" UUID REFERENCES public.workspaces(id) ON DELETE CASCADE,
    type TEXT NOT NULL DEFAULT 'comentario',
    content TEXT NOT NULL DEFAULT '',
    author TEXT NOT NULL DEFAULT '',
    "photo_urls" TEXT DEFAULT '',
    "createdAt" TIMESTAMPTZ DEFAULT NOW()
);
ALTER TABLE public.ticket_events ADD COLUMN IF NOT EXISTS "ticket_id" UUID;
ALTER TABLE public.ticket_events ADD COLUMN IF NOT EXISTS "workspace_id" UUID;
ALTER TABLE public.ticket_events ADD COLUMN IF NOT EXISTS type TEXT NOT NULL DEFAULT 'comentario';
ALTER TABLE public.ticket_events ADD COLUMN IF NOT EXISTS content TEXT NOT NULL DEFAULT '';
ALTER TABLE public.ticket_events ADD COLUMN IF NOT EXISTS author TEXT NOT NULL DEFAULT '';
ALTER TABLE public.ticket_events ADD COLUMN IF NOT EXISTS "photo_urls" TEXT DEFAULT '';
ALTER TABLE public.ticket_events ADD COLUMN IF NOT EXISTS "createdAt" TIMESTAMPTZ DEFAULT NOW();
ALTER TABLE public.chamados_tickets ADD COLUMN IF NOT EXISTS "priority" TEXT DEFAULT 'normal';
ALTER TABLE public.chamados_tickets ADD COLUMN IF NOT EXISTS "archived" BOOLEAN NOT NULL DEFAULT FALSE;
ALTER TABLE public.chamados_tickets ADD COLUMN IF NOT EXISTS "closedAt" TIMESTAMPTZ;
ALTER TABLE public.chamados_tickets ADD COLUMN IF NOT EXISTS "closedBy" TEXT DEFAULT '';
ALTER TABLE public.chamados_tickets ADD COLUMN IF NOT EXISTS "statusNote" TEXT DEFAULT '';
ALTER TABLE public.chamados_tickets ADD COLUMN IF NOT EXISTS "assignedToUserId" TEXT DEFAULT '';
ALTER TABLE public.chamados_tickets ADD COLUMN IF NOT EXISTS "photos" TEXT DEFAULT '';
ALTER TABLE public.chamados_tickets ADD COLUMN IF NOT EXISTS "feedbackRating" INTEGER;
ALTER TABLE public.chamados_tickets ADD COLUMN IF NOT EXISTS "feedbackComment" TEXT DEFAULT '';
ALTER TABLE public.chamados_tickets ADD COLUMN IF NOT EXISTS "feedbackAt" TIMESTAMPTZ;
DO $$ BEGIN
  ALTER TABLE public.chamados_tickets ADD CONSTRAINT chk_feedback_rating
    CHECK ("feedbackRating" IS NULL OR "feedbackRating" BETWEEN 1 AND 5);
EXCEPTION WHEN duplicate_object THEN NULL;
END $$;
CREATE INDEX IF NOT EXISTS idx_chamados_workspace ON public.chamados_tickets("workspace_id");
CREATE INDEX IF NOT EXISTS idx_chamados_status ON public.chamados_tickets(status);
CREATE INDEX IF NOT EXISTS idx_ticket_events_ticket ON public.ticket_events("ticket_id");
-- RLS: acesso direto (anon/authenticated) bloqueado; a API acessa via service role.
ALTER TABLE public.chamados_tickets ENABLE ROW LEVEL SECURITY;
REVOKE ALL ON public.chamados_tickets FROM anon, authenticated, PUBLIC;
ALTER TABLE public.ticket_events ENABLE ROW LEVEL SECURITY;
REVOKE ALL ON public.ticket_events FROM anon, authenticated, PUBLIC;
"""

CHAMADOS_PRIORITIES = ('baixa', 'normal', 'alta', 'urgente')
CHAMADOS_STATUSES = ('aberto', 'a_caminho', 'em_atendimento', 'resolvido', 'fechado')

CHAMADOS_STATUS_LABELS = {
    'aberto': 'Aguardando técnico',
    'a_caminho': 'Técnico a caminho',
    'em_atendimento': 'Atendendo agora',
    'resolvido': 'Chamado resolvido',
    'fechado': 'Chamado concluído',
}

CHAMADOS_SUBS_PER_TICKET = 10


def _chamado_subs(ticket_id):
    """Inscrições de push registradas para um chamado (professor, página pública)."""
    if not redis:
        return []
    raw = redis.smembers(f'push:chamado:{ticket_id}')
    subs = []
    for r in raw:
        try:
            subs.append(json.loads(r) if isinstance(r, str) else r)
        except Exception:
            continue
    return subs


def _save_chamado_subs(ticket_id, subs):
    if not redis:
        return
    key = f'push:chamado:{ticket_id}'
    redis.delete(key)
    for s in subs[:CHAMADOS_SUBS_PER_TICKET]:
        redis.sadd(key, json.dumps(s, ensure_ascii=False))


def _record_ticket_event(ticket_id, workspace_id, event_type, content='', author='', photo_urls=''):
    """Insere um evento no histórico do chamado (falhas não quebram o fluxo principal)."""
    if not _SUPABASE_URL or not _SUPABASE_SERVICE_KEY:
        return None
    try:
        payload = {
            'ticket_id': ticket_id,
            'workspace_id': workspace_id or None,
            'type': event_type,
            'content': content,
            'author': author,
            'photo_urls': photo_urls,
        }
        resp = requests.post(
            f'{_SUPABASE_URL}/rest/v1/ticket_events',
            headers={**_supabase_headers(), 'Prefer': 'return=representation'},
            json=payload,
            timeout=10,
        )
        if not resp.ok:
            print(f"[chamados] erro ao gravar evento: {resp.status_code} {resp.text[:200]}")
            return None
        rows = resp.json() or []
        return rows[0] if rows else None
    except Exception as e:
        logger.error("[chamados] evento error: %s", e)
        return None


def _notify_ticket_status(ticket):
    """Push ao professor (inscrições do próprio chamado) quando status/mensagem mudam."""
    try:
        subs = _chamado_subs(ticket.get('id', ''))
        if not subs:
            return
        status = ticket.get('status', '')
        note = (ticket.get('statusNote') or '').strip()

        if status == 'resolvido':
            title = 'Como foi seu atendimento? ⭐'
            body = f"O chamado #{ticket.get('ticketNumber')} foi resolvido. Avalie o atendimento da equipe de TI."
            url = f"/chamados-publico/feedback/{ticket.get('id')}"
        else:
            label = CHAMADOS_STATUS_LABELS.get(status, status)
            msg = f'{label} — {note}' if note else label
            title = f"Chamado #{ticket.get('ticketNumber')}: {msg}"
            body = ' · '.join(
                str(part) for part in (ticket.get('roomName'), ticket.get('problemCategory')) if part
            )
            url = f"/chamados-publico/success/{ticket.get('id')}"

        keep = []
        for sub in subs:
            if push_notify(sub, title, body, url=url):
                keep.append(sub)
        _save_chamado_subs(ticket.get('id', ''), keep)
        print(f"[chamados] push status #{ticket.get('ticketNumber')}: {len(keep)}/{len(subs)} enviados")
    except Exception as e:
        logger.error("[chamados] push status error: %s", e)


def _notify_ticket_assigned(ticket):
    """Push direto ao técnico atribuído quando o chamado é designado a ele.

    Segmenta pelas inscrições do usuário (user_id) com acesso ao módulo chamados
    no workspace do chamado. Falha de push não quebra o fluxo de atualização.
    """
    try:
        user_id = ticket.get('assignedToUserId') or ''
        if not user_id:
            return
        subs = _target_subs(
            module='chamados',
            workspace_id=ticket.get('workspace_id'),
            user_id=user_id,
        )
        if not subs:
            return
        title = f"Chamado #{ticket.get('ticketNumber')} atribuído a você"
        body = ' · '.join(
            str(part) for part in (ticket.get('roomName'), ticket.get('problemCategory'), ticket.get('reportedBy')) if part
        )
        url = f"/chamados/tickets/{ticket.get('id')}"
        sent = 0
        for sub in subs:
            if push_notify(sub, title, body, url=url):
                sent += 1
        print(f"[chamados] push atribuição #{ticket.get('ticketNumber')}: {sent}/{len(subs)} enviados")
    except Exception as e:
        logger.error("[chamados] push atribuição error: %s", e)


def _notify_ticket_claimed(ticket, claimer_name):
    """Notifica a assunção de um chamado.

    - Ao técnico que assumiu: "Você assumiu o chamado #N."
    - Aos demais técnicos do mesmo workspace: "O chamado #N foi assumido por X."

    Não cria tempestade de notificações: cada subscriber recebe exatamente uma
    das duas mensagens. Falhas de push não quebram o fluxo de claim.
    """
    try:
        claimer_id = str(ticket.get('assignedToUserId') or '').strip()
        if not claimer_id:
            return
        num = ticket.get('ticketNumber')
        url = f"/chamados/tickets/{ticket.get('id')}"
        ws = ticket.get('workspace_id')

        # 1. Push direto a quem assumiu.
        own_subs = _target_subs(module='chamados', workspace_id=ws, user_id=claimer_id)
        own_title = f"Você assumiu o chamado #{num}"
        own_body = ' · '.join(
            str(part) for part in (ticket.get('roomName'), ticket.get('problemCategory')) if part
        )
        sent_own = 0
        for sub in (own_subs or []):
            if push_notify(sub, own_title, own_body, url=url, user_id=claimer_id):
                sent_own += 1

        # 2. Demais técnicos do workspace: informação de indisponibilidade.
        others = _target_subs(module='chamados', workspace_id=ws)
        other_title = f"O chamado #{num} foi assumido por {claimer_name or 'um técnico'}"
        other_body = ' · '.join(
            str(part) for part in (ticket.get('roomName'), ticket.get('problemCategory')) if part
        )
        sent_other = 0
        for sub in (others or []):
            u = (sub.get('user') or {})
            if u.get('id') == claimer_id:
                continue
            if push_notify(sub, other_title, other_body, url=url):
                sent_other += 1
        print(
            f"[chamados] push claim #{num}: {sent_own} p/ responsável, {sent_other} p/ demais técnicos"
        )
    except Exception as e:
        logger.error("[chamados] push claim error: %s", e)


def _notify_new_ticket(ticket):
    """Dispara push imediato para o TI quando um chamado é aberto.

    Usa a mesma infra de push do app (Upstash Redis + VAPID + pywebpush):
    filtra os inscritos pelo módulo 'chamados' e pelo workspace do chamado,
    respeitando perfil (apps), cargo e notify_settings de cada usuário.
    Falha de push não impede a criação do chamado.
    """
    try:
        subs = _target_subs(module='chamados', workspace_id=ticket.get('workspace_id'))
        if not subs:
            return
        title = f"Novo chamado #{ticket.get('ticketNumber')}"
        body = ' · '.join(
            str(part) for part in (
                ticket.get('roomName'),
                ticket.get('problemCategory'),
                ticket.get('reportedBy'),
            ) if part
        )
        url = f"/chamados/tickets/{ticket.get('id')}"
        sent = 0
        for sub in subs:
            if push_notify(sub, title, body, url=url):
                sent += 1
        print(f"[chamados] push novo chamado #{ticket.get('ticketNumber')}: {sent}/{len(subs)} enviados")
    except Exception as e:
        logger.error("[chamados] push error: %s", e)


def _ensure_chamados_schema():
    """Cria a tabela de chamados se não existir (mesmo padrão do _ensure_stock_schema)."""
    if not _SUPABASE_URL or not _SUPABASE_SERVICE_KEY:
        return
    headers = {'apikey': _SUPABASE_SERVICE_KEY, 'Authorization': f'Bearer {_SUPABASE_SERVICE_KEY}'}
    try:
        resp = requests.post(
            f'{_SUPABASE_URL}/rest/v1/rpc/pg_sql',
            json={'query': CHAMADOS_TABLE_SQL},
            headers=headers,
            timeout=10,
        )
        print(f"[chamados] pg_sql: {resp.status_code} {resp.text[:200]}")
    except Exception as e:
        logger.error("[chamados] pg_sql error: %s", e)


@app.route('/api/chamados/workspaces', methods=['GET'])
def chamados_workspaces():
    """Lista os campi (workspaces) para o formulário público. Público, service role."""
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        _ensure_chamados_schema()
        resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/workspaces?select=id,name,slug,location&order=name',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not resp.ok:
            return jsonify({'error': 'Erro ao listar campi'}), 502
        return jsonify({'workspaces': resp.json() or []})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/chamados', methods=['POST'])
def chamados_create():
    """Cria um chamado a partir do formulário público (professor), sem login."""
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        # Rate limiting
        client_ip = _get_client_ip()
        if not _check_rate_limit(client_ip):
            return jsonify({'error': 'Muitas requisições. Tente novamente em mais de 1 hora.'}), 429

        _ensure_chamados_schema()
        body = request.get_json() or {}

        workspace_id = str(body.get('workspace_id') or '').strip()
        room_name = str(body.get('roomName') or '').strip()
        reported_by = str(body.get('reportedBy') or '').strip()
        problem_category = str(body.get('problemCategory') or '').strip()
        problem_area = str(body.get('problemArea') or '').strip()
        problem_description = str(body.get('problemDescription') or '').strip()
        priority = str(body.get('priority') or 'normal').strip().lower()

        if not workspace_id:
            return jsonify({'error': 'Selecione o campus'}), 400
        if not room_name:
            return jsonify({'error': 'Informe a sala'}), 400
        if not reported_by:
            return jsonify({'error': 'Informe seu nome'}), 400
        if problem_area not in ('administrativa', 'academica'):
            return jsonify({'error': 'Selecione a área do problema'}), 400
        if not problem_category:
            return jsonify({'error': 'Selecione o tipo de problema'}), 400
        if not problem_description:
            return jsonify({'error': 'Descreva o que está acontecendo'}), 400
        if priority not in CHAMADOS_PRIORITIES:
            return jsonify({'error': 'Prioridade inválida'}), 400

        # Validação de tamanho de texto
        field_lengths = {
            'workspace_id': workspace_id, 'roomName': room_name,
            'reportedBy': reported_by, 'problemCategory': problem_category,
            'problemArea': problem_area, 'problemDescription': problem_description,
        }
        for key, value in field_lengths.items():
            err = _validate_field_length(key, value)
            if err:
                return jsonify({'error': err}), 400

        photos = str(body.get('photos') or '').strip()
        if len(photos) > 600000:
            return jsonify({'error': 'Foto muito grande'}), 400
        if photos and not _is_valid_photo(photos):
            return jsonify({'error': 'Foto inválida'}), 400

        ws_check = requests.get(
            f'{_SUPABASE_URL}/rest/v1/workspaces?id=eq.{quote(workspace_id)}&select=id,disabled_apps',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not ws_check.ok or not ws_check.json():
            return jsonify({'error': 'Campus não encontrado'}), 400

        workspace = ws_check.json()[0]
        module_err = require_module(workspace, 'chamados')
        if module_err:
            return module_err

        num_resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/chamados_tickets'
            f'?select=ticketNumber&workspace_id=eq.{quote(workspace_id)}'
            f'&order=ticketNumber.desc&limit=1',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not num_resp.ok:
            return jsonify({'error': 'Erro ao gerar o número do chamado'}), 502
        num_rows = num_resp.json()
        ticket_number = (num_rows[0].get('ticketNumber') if num_rows else 0) + 1

        now = datetime.now(timezone.utc).isoformat()

        # Gera o tracking token (credential de escopo limitado do professor).
        # O token cru é entregue ao professor APENAS nesta resposta; somente o
        # hash SHA-256 é persistido no banco. Nunca logado nem devolvido de novo.
        tracking_token = secrets.token_urlsafe(32)
        tracking_token_hash = hashlib.sha256(tracking_token.encode('utf-8')).hexdigest()

        payload = {
            'workspace_id': workspace_id,
            'roomId': str(body.get('roomId') or '').strip(),
            'roomName': room_name,
            'assetId': str(body.get('assetId') or ''),
            'assetSource': str(body.get('assetSource') or ''),
            'assetName': str(body.get('assetName') or ''),
            'assetPatrimony': str(body.get('assetPatrimony') or ''),
            'problemCategory': problem_category,
            'problemArea': problem_area,
            'problemDescription': problem_description,
            'status': 'aberto',
            'priority': priority,
            'reportedBy': reported_by,
            'reportedByEmail': str(body.get('reportedByEmail') or '').strip(),
            'assignedTo': str(body.get('assignedTo') or ''),
            'assignedToUserId': str(body.get('assignedToUserId') or ''),
            'photos': photos,
            'ticketNumber': ticket_number,
            'tracking_token_hash': tracking_token_hash,
            'createdAt': now,
            'updatedAt': now,
            'resolvedAt': None,
        }

        ins = requests.post(
            f'{_SUPABASE_URL}/rest/v1/chamados_tickets',
            headers={**_supabase_headers(), 'Prefer': 'return=representation'},
            json=payload,
            timeout=10,
        )
        if not ins.ok:
            return jsonify({'error': f'Erro ao criar chamado: {ins.status_code} {ins.text[:200]}'}), 502

        ticket = ins.json()[0]

        # Validação de integridade: garante que campos obrigatórios foram persistidos.
        if not ticket.get('createdAt') or ticket.get('ticketNumber') is None:
            logger.error(
                "[chamados] Ticket criado sem campos obrigatórios: createdAt=%s ticketNumber=%s",
                ticket.get('createdAt'), ticket.get('ticketNumber'),
            )
            ticket.setdefault('createdAt', now)
            ticket.setdefault('ticketNumber', ticket_number)

        # Notificação imediata: avisa o TI no momento em que o professor abre o chamado.
        # Evento, não agendamento — não depende de cron nenhum (próprio app).
        _notify_new_ticket(ticket)

        # O token cru NUNCA entra na resposta persistida (ticket); vai apenas no
        # campo tracking_token desta única resposta, para o professor guardar.
        # O tracking_token_hash (SHA-256) também não é devolvido ao cliente:
        # é segredo interno usado apenas na autenticação por token.
        ticket.pop('tracking_token_hash', None)
        return jsonify({'ticket': ticket, 'tracking_token': tracking_token})

    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/chamados', methods=['GET'])
@require_auth
def chamados_list():
    """Lista chamados (filtros opcionais: workspace_id, status, reportedBy). Requer autenticação."""
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        _ensure_chamados_schema()
        user = g.user
        user_ws_ids = [str(w) for w in (user.get('workspace_ids') or [])]
        is_super_admin = bool(user.get('is_super_admin'))

        workspace_id = request.args.get('workspace_id')
        if workspace_id:
            if not is_super_admin and workspace_id not in user_ws_ids:
                return jsonify({'error': 'Acesso negado a este workspace'}), 403

        url = f'{_SUPABASE_URL}/rest/v1/chamados_tickets?select=*&order=createdAt.desc'
        if workspace_id:
            url += f'&workspace_id=eq.{quote(workspace_id)}'
        elif not is_super_admin:
            if not user_ws_ids:
                return jsonify({'tickets': []})
            if len(user_ws_ids) == 1:
                url += f'&workspace_id=eq.{quote(user_ws_ids[0])}'
            else:
                ws_filter = ','.join(f'"{w}"' for w in user_ws_ids)
                url += f'&workspace_id=in.({quote(ws_filter)})'

        status = request.args.get('status')
        if status:
            url += f'&status=eq.{quote(status)}'
        reported_by = request.args.get('reportedBy')
        if reported_by:
            url += f'&reportedBy=ilike.*{quote(reported_by)}*'
        resp = requests.get(url, headers=_supabase_headers(), timeout=15)
        if not resp.ok:
            return jsonify({'error': 'Erro ao listar chamados'}), 502
        return jsonify({'tickets': resp.json() or []})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/chamados/<ticket_id>', methods=['GET', 'PATCH', 'DELETE'])
@require_auth
def chamados_manage(ticket_id):
    """Consulta, atualiza status/responsável/prioridade de um chamado ou remove. Requer autenticação."""
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        _ensure_chamados_schema()
        user = g.user
        is_super_admin = bool(user.get('is_super_admin'))
        user_ws_ids = set(str(w) for w in (user.get('workspace_ids') or []))

        if request.method == 'GET':
            resp = requests.get(
                f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket_id)}&select=*',
                headers=_supabase_headers(),
                timeout=10,
            )
            if not resp.ok:
                return jsonify({'error': 'Erro ao buscar chamado'}), 502
            rows = resp.json() or []
            if not rows:
                return jsonify({'error': 'Chamado não encontrado'}), 404
            ticket_ws = rows[0].get('workspace_id') or ''
            if not is_super_admin and (not ticket_ws or ticket_ws not in user_ws_ids):
                return jsonify({'error': 'Acesso negado a este chamado'}), 403
            # Etapa 6 — workspace do recurso resolvido; enforce ticket.view.
            g.workspace_id = ticket_ws
            err = _require_action_in_handler('ticket.view', scope='workspace',
                                             resource_type='ticket', resource_id=ticket_id)
            if err:
                return err
            return jsonify({'ticket': rows[0]})

        if request.method == 'DELETE':
            # Verify workspace ownership before delete
            fetch = requests.get(
                f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket_id)}&select=workspace_id,photos',
                headers=_supabase_headers(),
                timeout=10,
            )
            if not fetch.ok:
                return jsonify({'error': 'Erro ao buscar chamado'}), 502
            rows = fetch.json() or []
            if not rows:
                return jsonify({'error': 'Chamado não encontrado'}), 404
            ticket_ws = rows[0].get('workspace_id') or ''
            if not is_super_admin and (not ticket_ws or ticket_ws not in user_ws_ids):
                return jsonify({'error': 'Acesso negado a este chamado'}), 403
            # Etapa 6 — workspace do recurso resolvido; enforce ticket.delete
            # ANTES de qualquer side effect (remoção do registro / Cloudinary).
            g.workspace_id = ticket_ws
            err = _require_action_in_handler('ticket.delete', scope='workspace',
                                             resource_type='ticket', resource_id=ticket_id)
            if err:
                return err

            # A1: coleta as fotos (Cloudinary) antes de apagar, para
            # limpeza best-effort após a remoção do registro.
            cloud_urls = []
            ticket_photo = str(rows[0].get('photos') or '').strip()
            if ticket_photo and ticket_photo.startswith('https://res.cloudinary.com/'):
                cloud_urls.append(ticket_photo)
            if rows[0].get('photos'):
                try:
                    ev_resp = requests.get(
                        f'{_SUPABASE_URL}/rest/v1/ticket_events?ticket_id=eq.{quote(ticket_id)}&select=photo_urls',
                        headers=_supabase_headers(),
                        timeout=10,
                    )
                    if ev_resp.ok:
                        for ev in (ev_resp.json() or []):
                            try:
                                urls = json.loads(ev.get('photo_urls') or '[]')
                            except (TypeError, ValueError):
                                urls = []
                            cloud_urls.extend(
                                u for u in urls
                                if isinstance(u, str) and u.startswith('https://res.cloudinary.com/')
                            )
                except Exception as e:
                    logger.warning('DELETE chamado: falha ao consultar eventos fotos: %s', e)

            resp = requests.delete(
                f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket_id)}',
                headers={**_supabase_headers(), 'Prefer': 'return=minimal'},
                timeout=10,
            )
            if not resp.ok:
                return jsonify({'error': 'Erro ao remover chamado'}), 502

            # Limpeza best-effort sequencial (nunca falha a resposta do DELETE).
            # O registro já foi removido; se a foto falhar, o purge diário é a 2ª camada
            # (porém não há mais ticket para achá-la — envio de log para rastreio).
            destroyed = 0
            for u in dict.fromkeys(cloud_urls):
                if _cloudinary_destroy(u):
                    destroyed += 1
                else:
                    logger.warning('DELETE chamado %s: falha ao apagar foto no Cloudinary: %s', ticket_id, u)
            return jsonify({'success': True, 'photos_destroyed': destroyed})

        # PATCH: Verify workspace ownership before any update
        fetch_ws = requests.get(
            f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket_id)}&select=workspace_id,status,assignedToUserId',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not fetch_ws.ok:
            return jsonify({'error': 'Erro ao buscar chamado'}), 502
        ws_rows = fetch_ws.json() or []
        if not ws_rows:
            return jsonify({'error': 'Chamado não encontrado'}), 404
        ticket_ws = ws_rows[0].get('workspace_id') or ''
        if not is_super_admin and (not ticket_ws or ticket_ws not in user_ws_ids):
            return jsonify({'error': 'Acesso negado a este chamado'}), 403
        # Etapa 6 — workspace do recurso resolvido (nunca o do cliente).
        g.workspace_id = ticket_ws

        # Ownership (sempre ativo): um técnico comum não pode operar um chamado
        # já atribuído a outro técnico. O responsável, o líder/assigner e o super
        # admin podem. Aplicado ANTES de qualquer mutation.
        _current_ticket = {
            'workspace_id': ticket_ws,
            'status': ws_rows[0].get('status'),
            'assignedToUserId': ws_rows[0].get('assignedToUserId') or '',
        }
        err = _enforce_ownership(user, _current_ticket, ticket_ws, ticket_id)
        if err:
            return err

        body = request.get_json() or {}
        updates = {}
        for key in ('status', 'assignedTo', 'assignedToUserId', 'problemDescription', 'priority', 'archived', 'closedAt', 'closedBy', 'statusNote', 'photos'):
            if key in body:
                updates[key] = body[key]
        if 'priority' in updates and updates['priority'] not in CHAMADOS_PRIORITIES:
            return jsonify({'error': 'Prioridade inválida'}), 400
        # Validação de tamanho de texto nos updates
        for key in ('assignedTo', 'assignedToUserId', 'problemDescription', 'statusNote', 'author'):
            if key in updates and isinstance(updates[key], str):
                err = _validate_field_length(key, updates[key])
                if err:
                    return jsonify({'error': err}), 400
        if 'photos' in updates:
            photos_val = str(updates['photos'] or '').strip()
            if len(photos_val) > 600000:
                return jsonify({'error': 'Foto muito grande'}), 400
            if photos_val and not _is_valid_photo(photos_val):
                return jsonify({'error': 'Foto inválida'}), 400
            # A2: captura a foto atual para, após o PATCH bem-sucedido, apagar
            # a foto antiga do Cloudinary (nunca apaga a nova).
            try:
                fetch_photos = requests.get(
                    f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket_id)}&select=photos',
                    headers=_supabase_headers(),
                    timeout=10,
                )
                if fetch_photos.ok:
                    _prev_photo = str((fetch_photos.json() or [{}])[0].get('photos') or '').strip()
                else:
                    _prev_photo = ''
            except Exception as e:
                logger.warning('PATCH chamado: falha ao consultar foto anterior: %s', e)
                _prev_photo = ''
            updates['photos'] = photos_val

        # Etapa 6 — PATCH mixed-operation: determinar o conjunto MÍNIMO de Actions
        # exigido pelas operações efetivamente solicitadas e autorizar TODAS
        # ANTES de qualquer mutation (atomicidade). Se qualquer Action for negada,
        # NADA é alterado (403).
        required_actions = set()
        if 'status' in updates or 'statusNote' in updates:
            required_actions.add('ticket.status')
        if 'assignedTo' in updates or 'assignedToUserId' in updates:
            required_actions.add('ticket.assign')
        if any(k in updates for k in (
            'problemDescription', 'priority', 'archived', 'closedAt', 'closedBy', 'photos'
        )):
            required_actions.add('ticket.edit')
        for act in sorted(required_actions):
            err = _require_action_in_handler(act, scope='workspace',
                                             resource_type='ticket', resource_id=ticket_id)
            if err:
                return err

        # Atribuição/reatribuição/remoção de responsável = privilégio de assigner.
        # O técnico comum NÃO pode escolher outro técnico, nem remover responsável,
        # nem (por esta rota) auto-atribuir-se — a auto-atribuição (claim) tem rota
        # dedicada e atômica. Bloqueamos inclusive com RBAC_2_ENABLED=0.
        if 'assignedTo' in updates or 'assignedToUserId' in updates:
            if not _is_assigner(user, ticket_ws):
                return _forbidden('Permissão insuficiente para atribuir responsável')

        prev = None
        assignment_changed = False
        if 'assignedToUserId' in updates:
            # Busca o responsável atual para só notificar quando houver troca/atribuição
            fetch = requests.get(
                f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket_id)}&select=assignedToUserId',
                headers=_supabase_headers(),
                timeout=10,
            )
            if not fetch.ok:
                return jsonify({'error': 'Erro ao buscar chamado'}), 502
            prev = (fetch.json() or [{}])[0]
            assignment_changed = (prev.get('assignedToUserId') or '') != (updates.get('assignedToUserId') or '')
        if 'status' in updates:
            status = updates['status']
            if status not in CHAMADOS_STATUSES:
                return jsonify({'error': 'Status inválido'}), 400

            # Busca o estado atual para detectar reabertura (resolvido/fechado → fluxo ativo)
            fetch = requests.get(
                f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket_id)}&select=status,statusNote,resolvedAt,closedAt,archived',
                headers=_supabase_headers(),
                timeout=10,
            )
            if not fetch.ok:
                return jsonify({'error': 'Erro ao buscar chamado'}), 502
            prev = (fetch.json() or [{}])[0]
            prev_status = prev.get('status')

            if status == 'resolvido':
                updates['resolvedAt'] = datetime.now(timezone.utc).isoformat()
                updates['statusNote'] = ''
            elif status == 'fechado':
                updates['archived'] = True
                updates['closedAt'] = datetime.now(timezone.utc).isoformat()
                updates['statusNote'] = ''
            elif prev_status in ('resolvido', 'fechado'):
                # Reabertura: volta ao fluxo ativo e limpa marcas de conclusão
                updates['resolvedAt'] = None
                updates['closedAt'] = None
                updates['closedBy'] = ''
                updates['archived'] = False
        elif 'statusNote' in updates:
            # Busca o estado atual para comparar a mensagem e notificar só em mudança
            fetch = requests.get(
                f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket_id)}&select=status,statusNote',
                headers=_supabase_headers(),
                timeout=10,
            )
            if not fetch.ok:
                return jsonify({'error': 'Erro ao buscar chamado'}), 502
            prev = (fetch.json() or [{}])[0]
        if not updates:
            return jsonify({'error': 'Nada para atualizar'}), 400
        updates['updatedAt'] = datetime.now(timezone.utc).isoformat()

        resp = requests.patch(
            f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket_id)}',
            headers={**_supabase_headers(), 'Prefer': 'return=representation'},
            json=updates,
            timeout=10,
        )
        if not resp.ok:
            return jsonify({'error': 'Erro ao atualizar chamado'}), 502
        rows = resp.json()
        if not rows:
            return jsonify({'error': 'Chamado não encontrado'}), 404

        ticket = rows[0]

        # A2: após PATCH bem-sucedido com troca de foto, apaga a foto antiga
        # do Cloudinary (best-effort, nunca bloqueia o PATCH, nunca apaga a nova).
        if 'photos' in updates and _prev_photo:
            new_photo = str(updates.get('photos') or '').strip()
            if (
                _prev_photo != new_photo
                and _prev_photo.startswith('https://res.cloudinary.com/')
                and _cloudinary_destroy(_prev_photo)
            ):
                pass
            elif _prev_photo != new_photo:
                logger.warning(
                    'PATCH chamado %s: falha ao apagar foto antiga no Cloudinary: %s',
                    ticket_id,
                    _prev_photo,
                )

        if assignment_changed and ticket.get('assignedToUserId'):
            # Push direto ao técnico atribuído (filtro por user_id no _target_subs).
            _notify_ticket_assigned(ticket)
        if prev is not None and ('status' in updates or 'statusNote' in updates):
            changed = ('status' in updates and prev.get('status') != ticket.get('status')) or (
                'statusNote' in updates and prev.get('statusNote') != ticket.get('statusNote')
            )
            if changed:
                _notify_ticket_status(ticket)
                note = str(body.get('statusNote') or ticket.get('statusNote') or '').strip()
                content = note or CHAMADOS_STATUS_LABELS.get(ticket.get('status', ''), ticket.get('status', ''))
                author = str(body.get('author') or '').strip() or 'Sistema'
                _record_ticket_event(
                    ticket_id,
                    ticket.get('workspace_id'),
                    'status',
                    content=content,
                    author=author,
                )
        return jsonify({'ticket': ticket})

    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


def _aggregate_ticket_reports(rows):
    """Agrega chamados do período em métricas para a página de relatórios."""
    total = len(rows)
    by_status = {}
    by_priority = {}
    by_category = {}
    by_area = {}
    by_room = {}
    tech = {}
    resolution_total_ms = 0
    resolution_count = 0
    rating_sum = 0
    rating_count = 0

    for t in rows:
        status = t.get('status') or 'aberto'
        by_status[status] = by_status.get(status, 0) + 1
        priority = t.get('priority') or 'normal'
        by_priority[priority] = by_priority.get(priority, 0) + 1
        category = t.get('problemCategory') or 'Outros'
        by_category[category] = by_category.get(category, 0) + 1
        area = t.get('problemArea') or ''
        if area:
            by_area[area] = by_area.get(area, 0) + 1
        room = t.get('roomName') or ''
        if room:
            by_room[room] = by_room.get(room, 0) + 1

        name = (t.get('assignedTo') or '').strip()
        if name:
            row = tech.setdefault(name, {'name': name, 'open': 0, 'resolved': 0, 'total': 0, 'resolutionMs': 0, 'ratingSum': 0, 'ratingCount': 0})
            row['total'] += 1
            if status in ('aberto', 'a_caminho', 'em_atendimento'):
                row['open'] += 1
            if status in ('resolvido', 'fechado') or t.get('resolvedAt'):
                row['resolved'] += 1
            if t.get('resolvedAt') and t.get('createdAt'):
                try:
                    row['resolutionMs'] += (_parse_iso(t['resolvedAt'], 'resolvedAt') - _parse_iso(t['createdAt'], 'createdAt')).total_seconds() * 1000
                except Exception:
                    pass
            rating = t.get('feedbackRating')
            if rating:
                row['ratingSum'] += float(rating)
                row['ratingCount'] += 1

        if t.get('resolvedAt') and t.get('createdAt'):
            try:
                ms = _parse_iso(t['resolvedAt'], 'resolvedAt') - _parse_iso(t['createdAt'], 'createdAt')
                resolution_total_ms += ms.total_seconds() * 1000
                resolution_count += 1
            except Exception:
                pass

        rating = t.get('feedbackRating')
        if rating:
            try:
                rating_sum += float(rating)
                rating_count += 1
            except (TypeError, ValueError):
                pass

    by_technician = []
    for row in tech.values():
        by_technician.append({
            'name': row['name'],
            'open': row['open'],
            'resolved': row['resolved'],
            'total': row['total'],
            'avgResolutionHours': round(row['resolutionMs'] / row['resolved'] / 3600000, 1) if row['resolved'] else None,
            'rating': round(row['ratingSum'] / row['ratingCount'], 2) if row['ratingCount'] else None,
            'ratingCount': row['ratingCount'],
        })
    by_technician.sort(key=lambda r: -r['total'])

    return {
        'total': total,
        'byStatus': by_status,
        'byPriority': by_priority,
        'byCategory': by_category,
        'byArea': by_area,
        'byRoom': sorted(by_room.items(), key=lambda kv: -kv[1]),
        'byTechnician': by_technician,
        'avgResolutionHours': round(resolution_total_ms / resolution_count / 3600000, 1) if resolution_count else None,
        'feedback': {'count': rating_count, 'average': round(rating_sum / rating_count, 2) if rating_count else None},
    }


def _parse_iso(value, label):
    try:
        s = value.replace('Z', '+00:00') if isinstance(value, str) else ''
        return datetime.fromisoformat(s)
    except (ValueError, TypeError):
        return None


@app.route('/api/chamados/reports', methods=['GET'])
@require_auth
def chamados_reports():
    """Relatório de chamados (agregação no servidor). Período via from/to (ISO).

    Métricas: total, por status/prioridade/categoria/área/sala, por técnico,
    tempo médio de resolução e satisfação. Fonte: Supabase (chamados_tickets).
    """
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        _ensure_chamados_schema()
        user = g.user
        is_super_admin = bool(user.get('is_super_admin'))
        user_ws_ids = [str(w) for w in (user.get('workspace_ids') or [])]

        now = datetime.now(timezone.utc)
        from_iso = request.args.get('from') or (now - timedelta(days=30)).isoformat()
        to_iso = request.args.get('to') or now.isoformat()
        from_dt = _parse_iso(from_iso, 'from')
        to_dt = _parse_iso(to_iso, 'to')
        if not from_dt or not to_dt:
            return jsonify({'error': 'Período inválido (use ISO, ex.: 2026-06-01T00:00:00Z)'}), 400
        if from_dt > to_dt:
            return jsonify({'error': 'from deve ser anterior a to'}), 400

        url = (
            f'{_SUPABASE_URL}/rest/v1/chamados_tickets'
            f'?select=status,priority,problemCategory,problemArea,roomName,assignedTo,createdAt,resolvedAt,feedbackRating'
            f'&createdAt=gte.{quote(from_iso)}&createdAt=lte.{quote(to_iso)}'
        )
        workspace_id = request.args.get('workspace_id')
        if workspace_id:
            if not is_super_admin and workspace_id not in user_ws_ids:
                return jsonify({'error': 'Acesso negado a este workspace'}), 403
            url += f'&workspace_id=eq.{quote(workspace_id)}'
        elif not is_super_admin:
            if not user_ws_ids:
                report = _aggregate_ticket_reports([])
                report['period'] = {'from': from_iso, 'to': to_iso}
                return jsonify({'report': report})
            if len(user_ws_ids) == 1:
                url += f'&workspace_id=eq.{quote(user_ws_ids[0])}'
            else:
                ws_filter = ','.join(f'"{w}"' for w in user_ws_ids)
                url += f'&workspace_id=in.({quote(ws_filter)})'

        resp = requests.get(url, headers=_supabase_headers(), timeout=15)
        if not resp.ok:
            return jsonify({'error': 'Erro ao gerar relatório'}), 502

        report = _aggregate_ticket_reports(resp.json() or [])
        report['period'] = {'from': from_iso, 'to': to_iso}
        return jsonify({'report': report})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


# ──────────────────────────────────────────────────────────────────────────────
# Endpoints públicos do chamado (acesso do professor via tracking token)
# ──────────────────────────────────────────────────────────────────────────────
#
# O professor NÃO é autenticado — ele prova posse do chamado através do
# tracking token (hash SHA-256 validado pelo decorator). O ticket_id é SEMPRE
# derivado do token em g.tracking_ticket; nunca confiamos em ticket_id enviado
# pelo cliente. RLS permanece fechado; estas rotas são a única porta pública.

def _project_public_ticket(t: dict) -> dict:
    """Projeção segura do chamado para o professor (sem campos internos)."""
    return {
        'id': t.get('id'),
        'ticketNumber': t.get('ticketNumber'),
        'status': t.get('status'),
        'roomName': t.get('roomName'),
        'problemCategory': t.get('problemCategory'),
        'problemArea': t.get('problemArea'),
        'problemDescription': t.get('problemDescription'),
        'reportedBy': t.get('reportedBy'),
        'photos': t.get('photos'),
        'feedbackRating': t.get('feedbackRating'),
        'createdAt': t.get('createdAt'),
        'updatedAt': t.get('updatedAt'),
        'closedAt': t.get('closedAt'),
    }


@app.route('/api/public/chamados/<tracking_token>', methods=['GET'])
@require_tracking_token
def public_chamados_detail(tracking_token):
    """Status + dados básicos do chamado, acessível apenas com o token do próprio chamado."""
    ticket = g.tracking_ticket
    resp = requests.get(
        f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket["id"])}&select=*',
        headers=_supabase_headers(),
        timeout=10,
    )
    if not resp.ok or not resp.json():
        return jsonify({'error': 'Chamado não encontrado'}), 404
    return jsonify({'ticket': _project_public_ticket(resp.json()[0])})


@app.route('/api/public/chamados/<tracking_token>/events', methods=['GET'])
@require_tracking_token
def public_chamados_events(tracking_token):
    """Timeline do chamado, acessível apenas com o token do próprio chamado."""
    ticket = g.tracking_ticket
    resp = requests.get(
        f'{_SUPABASE_URL}/rest/v1/ticket_events?ticket_id=eq.{quote(ticket["id"])}'
        f'&order=createdAt.desc&select=id,type,content,author,photo_urls,createdAt',
        headers=_supabase_headers(),
        timeout=10,
    )
    if not resp.ok:
        return jsonify({'error': 'Erro ao carregar o histórico'}), 502
    events = []
    for ev in (resp.json() or []):
        try:
            urls = json.loads(ev.get('photo_urls') or '[]')
        except (TypeError, ValueError):
            urls = []
        events.append({
            'id': ev.get('id'),
            'type': ev.get('type'),
            'content': ev.get('content'),
            'author': ev.get('author'),
            'photos': urls if isinstance(urls, list) else [],
            'createdAt': ev.get('createdAt'),
        })
    return jsonify({'events': events})


@app.route('/api/public/chamados/<tracking_token>/feedback', methods=['POST'])
@require_tracking_token
def public_chamados_feedback(tracking_token):
    """Registra feedback (1-5) do professor para o próprio chamado.
    - Só permite quando resolvido/fechado.
    - Uma única vez por chamado (segunda tentativa → 409).
    - O ticket é derivado do token, nunca do corpo da requisição.
    """
    ticket = g.tracking_ticket
    fetch = requests.get(
        f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket["id"])}&select=status,feedbackRating',
        headers=_supabase_headers(),
        timeout=10,
    )
    if not fetch.ok or not fetch.json():
        return jsonify({'error': 'Chamado não encontrado'}), 404
    full = fetch.json()[0]

    if full.get('status') not in ('resolvido', 'fechado'):
        return jsonify({'error': 'Só é possível avaliar após a resolução do chamado'}), 403
    if full.get('feedbackRating') is not None:
        return jsonify({'error': 'Chamado já avaliado'}), 409

    body = request.get_json() or {}
    raw = body.get('rating')
    if isinstance(raw, float) and raw != int(raw):
        rating = 0
    else:
        try:
            rating = int(raw)
        except (TypeError, ValueError):
            rating = 0
    if rating not in (1, 2, 3, 4, 5):
        return jsonify({'error': 'Nota inválida (1 a 5)'}), 400
    comment = str(body.get('comment') or '').strip()[:500]

    now = datetime.now(timezone.utc).isoformat()
    resp = requests.patch(
        f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket["id"])}',
        headers={**_supabase_headers(), 'Prefer': 'return=representation'},
        json={
            'feedbackRating': rating,
            'feedbackComment': comment,
            'feedbackAt': now,
            'updatedAt': now,
        },
        timeout=10,
    )
    if not resp.ok or not resp.json():
        return jsonify({'error': 'Erro ao registrar o feedback'}), 502
    return jsonify({'ticket': _project_public_ticket(resp.json()[0])})


@app.route('/api/public/chamados/<tracking_token>/subscribe', methods=['POST'])
@require_tracking_token
def public_chamados_subscribe(tracking_token):
    """Registra push do professor para o próprio chamado (escopo limitado ao token)."""
    if not redis:
        return jsonify({'error': 'Redis não configurado'}), 500
    ticket = g.tracking_ticket
    body = request.get_json() or {}
    endpoint = body.get('endpoint', '')
    if not endpoint:
        return jsonify({'error': 'endpoint é obrigatório'}), 400

    sub = {
        'key': hashlib.sha256(endpoint.encode()).hexdigest(),
        'endpoint': endpoint,
        'expirationTime': body.get('expirationTime'),
        'keys': body.get('keys') or {},
    }
    subs = [s for s in _chamado_subs(ticket['id']) if s.get('endpoint') != endpoint]
    subs.append(sub)
    _save_chamado_subs(ticket['id'], subs)
    return jsonify({'status': 'ok', 'count': len(subs)})


@app.route('/api/chamados/push/test', methods=['POST'])
@require_auth
@require_admin
def chamados_push_test():
    """Envia uma push de teste para o próprio usuário logado (módulo chamados)."""
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        user_id = g.user_id

        subs = _target_subs(module='chamados', user_id=user_id)
        if not subs:
            return jsonify({
                'sent': 0,
                'total': 0,
                'message': 'Nenhuma inscrição push encontrada para este usuário. Ative as notificações primeiro.',
            })

        title = 'Teste de notificação — Chamados'
        body = 'Push funcionando! Você receberá avisos de novos chamados. 🔔'
        sent = 0
        for sub in subs:
            if push_notify(sub, title, body, url='/chamados'):
                sent += 1
        print(f"[chamados] push de teste: {sent}/{len(subs)} enviados (user={user_id})")
        return jsonify({'sent': sent, 'total': len(subs)})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


# ── Zerar dados (factory reset) ──

WIPE_TOKEN = os.environ.get('WIPE_TOKEN', '')

# Tabelas operacionais apagadas no wipe. Config (workspaces/perfis/cargos) é preservada.
WIPE_TABLES = [
    # public (filhos primeiro para respeitar FKs)
    ('public', 'tv_music_tracks'),
    ('public', 'tv_music_queues'),
    ('public', 'tv_gallery_photos'),
    ('public', 'tv_galleries'),
    ('public', 'tv_events'),
    ('public', 'tv_playlists'),
    ('public', 'tv_announcements'),
    ('public', 'tv_calendar_cache'),
    ('public', 'tv_urgent_announcements'),
    ('public', 'tv_activation_codes'),
    ('public', 'tablet_reservations'),
    ('public', 'chamados_tickets'),
    # schema stock
    ('stock', 'stock_movements'),
    ('stock', 'stock_maintenance'),
    ('stock', 'inventory_counts'),
    ('stock', 'inventory_cycles'),
    ('stock', 'stock_kits'),
    ('stock', 'notifications'),
    ('stock', 'stock_items'),
    # schema pcare
    ('pcare', 'action_logs'),
    ('pcare', 'pc_checklists'),
    ('pcare', 'checklist_templates'),
    ('pcare', 'part_usage'),
    ('pcare', 'maintenance'),
    ('pcare', 'parts'),
    ('pcare', 'pcs'),
]


@app.route('/api/admin/wipe', methods=['POST'])
@require_auth
@require_admin
@require_action_rbac('admin.system.wipe', scope='global')
def admin_wipe():
    """Apaga TODAS as linhas das tabelas operacionais (stock, pcare, chamados, TV).

    Requer o header `X-Wipe-Token` igual a WIPE_TOKEN (variável de ambiente).
    Auth required, super admin required.
    """
    token = (request.headers.get('X-Wipe-Token') or '').strip()
    if not WIPE_TOKEN or token != WIPE_TOKEN:
        return jsonify({'error': 'Token inválido ou não configurado'}), 403
    if not _SUPABASE_URL or not _SUPABASE_SERVICE_KEY:
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        results = {}
        for schema, table in WIPE_TABLES:
            headers = dict(_supabase_headers())
            if schema in ('stock', 'pcare'):
                headers['Accept-Profile'] = schema
                headers['Content-Profile'] = schema
            try:
                # WHERE é obrigatório (extensão safeupdate no Supabase).
                # 404 = tabela ainda não criada no projeto (sem dados para apagar).
                resp = requests.delete(
                    f'{_SUPABASE_URL}/rest/v1/{table}?id=neq.00000000-0000-0000-0000-000000000000',
                    headers={**headers, 'Prefer': 'return=minimal'},
                    timeout=20,
                )
                if resp.status_code == 204:
                    results[table] = 'ok'
                elif resp.status_code == 404:
                    results[table] = 'sem-tabela'
                else:
                    results[table] = f'HTTP {resp.status_code}'
            except Exception as e:
                logger.error("Erro ao limpar %s: %s", table, e)
                results[table] = 'erro'
        return jsonify({'wipe': results})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


# ── App data purge (TV) ──────────────────────────────────────────────────────
#
# Fluxo do PR 5: describe (contagens reais) → backup obrigatório → purge
# transacional → audit. Toda a lógica destrutiva vive em UMA função SQL
# (public.purge_tv_app_data, migration 032) executada via rpc: ou tudo
# acontece (backup + deletes + audit) ou nada acontece.
#
# Autorização: @require_auth + @require_workspace validam identidade e
# membership; o gate de gerência espelha can_manage_workspace_apps (migration
# 031): super admin OU membro com profile.role='admin'. O decorator existente
# require_admin é super-admin-only (plataforma) e não pode ser aplicado sem
# quebrar o requisito "admin do workspace pode purgar".
#
# Workspace NUNCA vem do corpo como autoridade: g.workspace_id é resolvido e
# validado pelo require_workspace (membership contra o JWT); o valor é apenas
# repassado às funções SQL, que escopam todos os predicates por ele.

APP_DATA_PURGE_MAX_ROWS = 50000
APP_DATA_PURGE_MAX_BYTES = 32 * 1024 * 1024  # 32 MB de JSONB

SUPPORTED_PURGE_APPS = ('tv',)


def _require_workspace_app_manager():
    """403 a menos que o usuário tenha permissão de gerência do app do workspace.

    Alinhado à autoridade atual do RBAC 2.0 (Etapa 3):
    - RBAC ativo ⇒ resolve 'admin.app.purge' no workspace via rbac_can
      (super admin ⇒ allow; permissões/overrides das tabelas RBAC; default deny).
    - Legacy (flag off) ⇒ espelho de can_manage_workspace_apps (migration 031):
      super admin OU membro com profile.role='admin' (compat preservada apenas
      durante a migração; profile.role NÃO é autoridade do RBAC 2.0).

    Retorna Response de erro ou None.
    """
    user = getattr(g, 'user', None) or {}
    if rbac_two_enabled():
        result = rbac_two_can(user, getattr(g, 'workspace_id', None), 'admin.app.purge', scope='workspace')
        if result:
            return None
        return _forbidden('Permissão insuficiente')
    if user.get('is_super_admin'):
        return None
    role = str(user.get('role') or '').strip().lower()
    if role != 'admin':
        return _forbidden('Workspace admin access required')
    return None


def _validate_purge_app_id():
    """Valida appId/app_id do corpo contra os apps suportados hoje (apenas tv).

    Retorna (app_id, error_response). Campos workspace_id/workspace no corpo são
    ignorados como autoridade — quem manda é g.workspace_id.
    """
    body = request.get_json(silent=True) or {}
    app_id = str(body.get('appId') or body.get('app_id') or '').strip()
    if not app_id:
        return None, (jsonify({'error': 'appId é obrigatório'}), 400)
    if app_id not in SUPPORTED_PURGE_APPS:
        return None, (jsonify({'error': f'App "{app_id}" não suporta limpeza de dados'}), 400)
    if body.get('workspaceId') and str(body['workspaceId']) != str(g.workspace_id):
        # Rejeita explicitamente tentativa de apontar outro workspace.
        return None, _forbidden('Access denied to this workspace')
    return app_id, None


def _rpc(function_name: str, params: dict):
    """Chama uma função SQL via PostgREST RPC com service_role."""
    return requests.post(
        f'{_SUPABASE_URL}/rest/v1/rpc/{function_name}',
        headers=_supabase_headers(),
        json=params,
        timeout=60,
    )


def _supabase_unavailable():
    return jsonify({'error': 'Supabase não configurado'}), 503


@app.route('/api/admin/app-data/describe', methods=['POST'])
@require_auth
@require_workspace
@require_module_auth('tv')
@require_action_rbac('admin.app.purge', scope='workspace')
def admin_app_data_describe():
    """Contagens reais dos dados de conteúdo da TV do workspace autenticado.

    Body: {"appId": "tv"} (obrigatório; outros apps → 400).
    Resposta: { ok, appId, workspaceId, tables:{...}, total }
    """
    manager_err = _require_workspace_app_manager()
    if manager_err:
        return manager_err

    if not _SUPABASE_URL or not _SUPABASE_SERVICE_KEY:
        return _supabase_unavailable()

    app_id, err = _validate_purge_app_id()
    if err:
        return err

    try:
        resp = _rpc('describe_tv_app_data', {'p_workspace': g.workspace_id})
        if resp.status_code != 200:
            logger.error("[app-data] describe rpc %s: %s", resp.status_code, resp.text[:300])
            return jsonify({'error': 'Não foi possível calcular as contagens'}), 502
        data = resp.json() or {}
        tables = data.get('tables') or {}
        return jsonify({
            'ok': True,
            'appId': app_id,
            'workspaceId': g.workspace_id,
            'tables': tables,
            'total': data.get('total', 0),
        })
    except Exception as e:
        logger.error("[app-data] describe error: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/admin/app-data/purge', methods=['POST'])
@require_auth
@require_workspace
@require_module_auth('tv')
@require_action_rbac('admin.app.purge', scope='workspace')
def admin_app_data_purge():
    """Limpa os dados de conteúdo da TV do workspace autenticado.

    Ordem garantida server-side (uma única transação SQL):
      contagens → guarda de tamanho → BACKUP → DELETEs → AUDIT.
    Backup falhou/estourou limite ⇒ nada é apagado.

    Body: {"appId": "tv"}. workspace_id do corpo nunca é autoridade.
    Concorrência: advisory lock por workspace dentro da função SQL serializa
    purges simultâneos (o segundo executa depois e encontra zero linhas).
    """
    manager_err = _require_workspace_app_manager()
    if manager_err:
        return manager_err

    if not _SUPABASE_URL or not _SUPABASE_SERVICE_KEY:
        return _supabase_unavailable()

    app_id, err = _validate_purge_app_id()
    if err:
        return err

    actor = g.user or {}
    actor_name = actor.get('name') or actor.get('email') or ''

    try:
        resp = _rpc('purge_tv_app_data', {
            'p_workspace': g.workspace_id,
            'p_actor_id': g.user_id,
            'p_actor_name': actor_name,
            'p_max_rows': APP_DATA_PURGE_MAX_ROWS,
            'p_max_bytes': APP_DATA_PURGE_MAX_BYTES,
        })

        if resp.status_code == 200:
            data = resp.json() or {}
            result_kind = data.get('result')
            deleted = data.get('deleted') or {}
            total = int(data.get('totalDeleted') or 0)
            payload = {
                'ok': True,
                'appId': app_id,
                'backupId': data.get('backupId'),
                'backupExpiresAt': data.get('backupExpiresAt'),
                'deleted': deleted,
                'totalDeleted': total,
                'auditId': data.get('auditId'),
            }
            if result_kind == 'empty':
                payload['empty'] = True
            logger.info(
                "[app-data] purge ws=%s app=%s total=%s backup=%s audit=%s actor=%s",
                g.workspace_id, app_id, total, data.get('backupId'), data.get('auditId'), g.user_id,
            )
            return jsonify(payload)

        text = resp.text or ''
        logger.error("[app-data] purge rpc %s: %s", resp.status_code, text[:300])
        if 'APP_DATA_BACKUP_TOO_LARGE' in text:
            return jsonify({
                'error': (
                    'O volume de dados excede o limite seguro de backup. '
                    'Nada foi removido. Solicite limpeza administrativa específica.'
                ),
                'code': 'backup_too_large',
            }), 413
        if 'WORKSPACE_NOT_FOUND' in text:
            return jsonify({'error': 'Workspace não encontrado'}), 404
        # Qualquer outro erro ⇒ rollback total no banco; mensagem segura.
        return jsonify({'error': 'Não foi possível concluir a limpeza. Nenhum dado foi removido.'}), 500
    except Exception as e:
        logger.error("[app-data] purge error: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/chamados/<ticket_id>/events', methods=['GET'])
@require_auth
def chamados_events_list(ticket_id):
    """Histórico (timeline) de um chamado, do mais novo para o mais antigo."""
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        _ensure_chamados_schema()
        user = g.user
        is_super_admin = bool(user.get('is_super_admin'))
        user_ws_ids = set(str(w) for w in (user.get('workspace_ids') or []))

        # Verify ticket belongs to user's workspace
        fetch_ticket = requests.get(
            f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket_id)}&select=workspace_id',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not fetch_ticket.ok:
            return jsonify({'error': 'Erro ao buscar chamado'}), 502
        t_rows = fetch_ticket.json() or []
        if not t_rows:
            return jsonify({'error': 'Chamado não encontrado'}), 404
        ticket_ws = t_rows[0].get('workspace_id') or ''
        if not is_super_admin and (not ticket_ws or ticket_ws not in user_ws_ids):
            return jsonify({'error': 'Acesso negado a este chamado'}), 403
        # Etapa 6 — workspace do recurso resolvido; enforce ticket.view (timeline).
        g.workspace_id = ticket_ws
        err = _require_action_in_handler('ticket.view', scope='workspace',
                                         resource_type='ticket', resource_id=ticket_id)
        if err:
            return err

        resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/ticket_events?ticket_id=eq.{quote(ticket_id)}&order=createdAt.desc',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not resp.ok:
            return jsonify({'error': 'Erro ao buscar o histórico'}), 502
        events = []
        for ev in (resp.json() or []):
            try:
                urls = json.loads(ev.get('photo_urls') or '[]')
            except (TypeError, ValueError):
                urls = []
            ev['photos'] = urls if isinstance(urls, list) else []
            events.append(ev)
        return jsonify({'events': events})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/chamados/<ticket_id>/events', methods=['POST'])
@require_auth
def chamados_events_create(ticket_id):
    """Comentário/suporte visual no chamado (solicitante ou técnico), com até 2 fotos."""
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        _ensure_chamados_schema()
        fetch = requests.get(
            f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket_id)}&select=id,workspace_id,status,assignedToUserId',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not fetch.ok:
            return jsonify({'error': 'Erro ao buscar chamado'}), 502
        rows = fetch.json() or []
        if not rows:
            return jsonify({'error': 'Chamado não encontrado'}), 404
        ticket = rows[0]

        # SEC-01: Verify ticket belongs to user's workspace
        user = g.user
        is_super_admin = bool(user.get('is_super_admin'))
        user_ws_ids = set(str(w) for w in (user.get('workspace_ids') or []))
        ticket_ws = ticket.get('workspace_id') or ''
        if not is_super_admin and (not ticket_ws or ticket_ws not in user_ws_ids):
            return jsonify({'error': 'Acesso negado a este chamado'}), 403
        # Etapa 6 — workspace do recurso resolvido; enforce ticket.comment ANTES
        # de criar o evento (timeline auditable).
        g.workspace_id = ticket_ws
        err = _require_action_in_handler('ticket.comment', scope='workspace',
                                         resource_type='ticket', resource_id=ticket_id)
        if err:
            return err

        # Ownership (sempre ativo): outro técnico não pode comentar/operar um
        # chamado que já pertence a outro técnico. Líder/admin/super podem.
        err = _enforce_ownership(user, ticket, ticket_ws, ticket_id)
        if err:
            return err

        body = request.get_json() or {}
        content = str(body.get('content') or '').strip()[:1000]
        author = str(body.get('author') or '').strip()[:120]
        photos = body.get('photos') or []

        if not isinstance(photos, list):
            return jsonify({'error': 'photos deve ser uma lista'}), 400
        if len(photos) > 2:
            return jsonify({'error': 'Máximo de 2 fotos por evento'}), 400
        for p in photos:
            pv = str(p or '').strip()
            if not _is_valid_photo(pv):
                return jsonify({'error': 'Foto inválida'}), 400
        if not content and not photos:
            return jsonify({'error': 'Escreva um comentário ou anexe uma foto'}), 400

        event = _record_ticket_event(
            ticket_id,
            ticket.get('workspace_id'),
            'comentario',
            content=content,
            author=author or 'Anônimo',
            photo_urls=json.dumps([str(p) for p in photos], ensure_ascii=False),
        )
        if not event:
            return jsonify({'error': 'Erro ao gravar o evento'}), 502
        try:
            event['photos'] = json.loads(event.get('photo_urls') or '[]')
        except (TypeError, ValueError):
            event['photos'] = []
        return jsonify({'event': event}), 201

    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/chamados/<ticket_id>/claim', methods=['POST'])
@require_auth
def chamados_claim(ticket_id):
    """COMEÇAR ATENDIMENTO — técnico assume o chamado para si.

    Modelo definitivo de atribuição:
      - Um técnico comum SÓ pode assumir um chamado SEM responsável.
      - A assunção é ATÔMICA no banco: a atualização só alcança linhas com
        assignedToUserId IS NULL. Se outro técnico assumiu primeiro, a atualização
        afeta 0 linhas e este request recebe 409 (já assumido).
      - Não altera o status (o responsável segue o fluxo de status depois).
      - Registra evento de atribuição + auditoria + notificações.

    Autorização:
      - Precisa ser membro do workspace do chamado.
      - RBAC ON  → exige Action `ticket.claim` (técnico tem; líder/super passam).
      - RBAC OFF → qualquer membro autenticado do workspace pode claim (legado).
      - Apenas o responsável, o líder/assigner ou o super admin podem claim de um
        chamado já atribuído (geralmente desnecessário, mas a rota rejeita se não).
    """
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        user = g.user
        is_super_admin = bool(user.get('is_super_admin'))
        user_ws_ids = set(str(w) for w in (user.get('workspace_ids') or []))

        fetch = requests.get(
            f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(ticket_id)}&select=id,workspace_id,status,assignedTo,assignedToUserId,ticketNumber,roomName,problemCategory',
            headers=_supabase_headers(),
            timeout=10,
        )
        if not fetch.ok:
            return jsonify({'error': 'Erro ao buscar chamado'}), 502
        rows = fetch.json() or []
        if not rows:
            return jsonify({'error': 'Chamado não encontrado'}), 404
        ticket = rows[0]
        ticket_ws = ticket.get('workspace_id') or ''
        if not is_super_admin and (not ticket_ws or ticket_ws not in user_ws_ids):
            return jsonify({'error': 'Acesso negado a este chamado'}), 403

        g.workspace_id = ticket_ws

        # RBAC ON → Action `ticket.claim`. OFF → no-op (legado).
        err = _require_action_in_handler('ticket.claim', scope='workspace',
                                         resource_type='ticket', resource_id=ticket_id)
        if err:
            return err

        # Ownership: se já tem responsável, só o próprio responsável, o
        # líder/assigner ou o super admin podem (re)assumir.
        err = _enforce_ownership(user, ticket, ticket_ws, ticket_id)
        if err:
            return err

        claimer_id = str(user.get('id') or '').strip()
        claimer_name = str(user.get('name') or '').strip() or 'Técnico'
        if not claimer_id:
            return jsonify({'error': 'Usuário inválido'}), 400

        # UPDATE ATÔMICO: só alcança linhas SEM responsável. 0 linhas ⇒ já assumido.
        now = datetime.now(timezone.utc).isoformat()
        upd_resp = requests.patch(
            f'{_SUPABASE_URL}/rest/v1/chamados_tickets'
            f'?id=eq.{quote(ticket_id)}&assignedToUserId=is.null',
            headers={**_supabase_headers(), 'Prefer': 'return=representation'},
            json={
                'assignedToUserId': claimer_id,
                'assignedTo': claimer_name,
                'updatedAt': now,
            },
            timeout=10,
        )
        if not upd_resp.ok:
            return jsonify({'error': 'Erro ao assumir chamado'}), 502
        updated_rows = upd_resp.json() or []
        if not updated_rows:
            # Chamado já foi assumido por outro técnico (race perdida).
            return jsonify({'error': 'Este chamado já foi assumido por outro técnico'}), 409

        updated = updated_rows[0]

        # Evento de atribuição no histórico.
        _record_ticket_event(
            ticket_id,
            ticket_ws,
            'atribuicao',
            content=f'{claimer_name} iniciou o atendimento',
            author=claimer_name,
        )

        # Auditoria do claim (side-channel RBAC).
        rbac_record_audit(
            actor_id=claimer_id,
            actor_is_super=user.get('is_super_admin'),
            action='ticket.claim',
            workspace_id=ticket_ws,
            scope='workspace',
            effect='allow',
            outcome='success',
            resource_type='ticket',
            resource_id=ticket_id,
            meta={'prev_owner': str(ticket.get('assignedToUserId') or ''), 'new_owner': claimer_id},
        )

        # Notifica quem assumiu e os demais técnicos do workspace.
        _notify_ticket_claimed(updated, claimer_name)

        return jsonify({'ticket': updated})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


def _send_email_via_resend(to, subject, html):
    """Envia email via Resend (API REST). Retorna (ok, erro)."""
    api_key = os.environ.get('RESEND_API_KEY', '').strip()
    from_addr = os.environ.get('EMAIL_FROM', '').strip() or 'LabHub <labhub@resend.dev>'
    if not api_key:
        return False, 'RESEND_API_KEY não configurado no servidor'
    if not to:
        return False, 'Informe o destinatário do email'
    try:
        resp = requests.post(
            'https://api.resend.com/emails',
            headers={'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'},
            json={'from': from_addr, 'to': [to], 'subject': subject, 'html': html},
            timeout=15,
        )
        if not resp.ok:
            return False, f'Resend retornou {resp.status_code}: {resp.text[:200]}'
        return True, ''
    except Exception as e:
        logger.error('Erro ao enviar email via Resend: %s', e)
        return False, 'Falha de rede ao enviar o email'


def _build_weekly_report_html(report, workspace_name):
    """Monta o HTML do resumo semanal de chamados (usado no email)."""
    total = report.get('total', 0)
    by_status = report.get('byStatus', {})
    by_room = report.get('byRoom', []) or []
    by_tech = report.get('byTechnician', []) or []
    feedback = report.get('feedback', {})

    status_rows = ''.join(
        f'<tr><td style="padding:6px 12px;border-bottom:1px solid #e5e7eb;color:#374151">{label}</td>'
        f'<td style="padding:6px 12px;border-bottom:1px solid #e5e7eb;color:#111827;text-align:right;font-weight:600">{by_status.get(status, 0)}</td></tr>'
        for status, label in [
            ('aberto', 'Abertos'),
            ('a_caminho', 'Técnico a caminho'),
            ('em_atendimento', 'Em atendimento'),
            ('resolvido', 'Resolvidos'),
            ('fechado', 'Fechados'),
        ]
    )

    room_rows = ''.join(
        f'<tr><td style="padding:5px 12px;border-bottom:1px solid #e5e7eb;color:#374151">{room}</td>'
        f'<td style="padding:5px 12px;border-bottom:1px solid #e5e7eb;color:#111827;text-align:right;font-weight:600">{count}</td></tr>'
        for room, count in by_room[:8]
    ) or '<tr><td style="padding:6px 12px;color:#6b7280">Sem chamados no período</td></tr>'

    tech_rows = ''.join(
        f'<tr><td style="padding:5px 12px;border-bottom:1px solid #e5e7eb;color:#374151">{t["name"]}</td>'
        f'<td style="padding:5px 12px;border-bottom:1px solid #e5e7eb;color:#111827;text-align:right">{t["resolved"]} resolvido(s)</td></tr>'
        for t in by_tech[:6]
    ) or '<tr><td style="padding:6px 12px;color:#6b7280">Nenhum técnico com chamados no período</td></tr>'

    avg = report.get('avgResolutionHours')
    avg_txt = f'{avg}h' if avg is not None else '—'
    fb_count = feedback.get('count', 0)
    fb_avg = feedback.get('average')
    fb_txt = f'{fb_avg:.1f} / 5 ({fb_count} avaliações)' if fb_avg is not None else f'{fb_count} avaliações' if fb_count else 'Sem avaliações'

    ws_name = workspace_name or 'todas as unidades'

    return f"""\
<!DOCTYPE html>
<html lang="pt-BR">
<body style="margin:0;padding:0;background:#f3f4f6;font-family:Arial,Helvetica,sans-serif">
  <div style="max-width:560px;margin:24px auto;background:#ffffff;border-radius:12px;overflow:hidden">
    <div style="background:linear-gradient(135deg,#f59e0b,#ea580c);padding:24px 28px">
      <h1 style="margin:0;color:#ffffff;font-size:20px">LabHub · Resumo Semanal</h1>
      <p style="margin:6px 0 0;color:#fef3c7;font-size:13px">{ws_name}</p>
    </div>
    <div style="padding:24px 28px">
      <table style="width:100%;border-collapse:collapse;margin-bottom:24px">
        <tr>
          <td style="text-align:center;background:#fef3c7;border-radius:8px;padding:14px">
            <p style="margin:0;font-size:28px;font-weight:700;color:#111827">{total}</p>
            <p style="margin:2px 0 0;font-size:12px;color:#6b7280">chamados na semana</p>
          </td>
          <td style="width:12px"></td>
          <td style="text-align:center;background:#ecfdf5;border-radius:8px;padding:14px">
            <p style="margin:0;font-size:28px;font-weight:700;color:#111827">{avg_txt}</p>
            <p style="margin:2px 0 0;font-size:12px;color:#6b7280">tempo médio de resolução</p>
          </td>
        </tr>
      </table>

      <h2 style="margin:0 0 8px;font-size:15px;color:#111827">Por status</h2>
      <table style="width:100%;border-collapse:collapse;margin-bottom:24px">{status_rows}</table>

      <h2 style="margin:0 0 8px;font-size:15px;color:#111827">Top salas</h2>
      <table style="width:100%;border-collapse:collapse;margin-bottom:24px">{room_rows}</table>

      <h2 style="margin:0 0 8px;font-size:15px;color:#111827">Técnicos</h2>
      <table style="width:100%;border-collapse:collapse;margin-bottom:24px">{tech_rows}</table>

      <p style="margin:0;font-size:13px;color:#374151">
        Satisfação dos professores: <strong>{fb_txt}</strong>
      </p>
    </div>
    <div style="padding:14px 28px;background:#f9fafb;border-top:1px solid #e5e7eb">
      <p style="margin:0;font-size:11px;color:#9ca3af">Gerado automaticamente pelo LabHub · {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')}</p>
    </div>
  </div>
</body>
</html>
"""


@app.route('/api/chamados/reports/weekly-email', methods=['POST'])
@require_auth
@require_admin
@require_action_rbac('ticket.weeklyEmail', scope='global')
def chamados_reports_weekly_email():
    """Envia por email o resumo semanal de chamados (últimos 7 dias)."""
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        _ensure_chamados_schema()
        now = datetime.now(timezone.utc)
        from_iso = (now - timedelta(days=7)).isoformat()
        to_iso = now.isoformat()

        body = request.get_json() or {}
        workspace_id = str(body.get('workspace_id') or '').strip()
        workspace_name = ''

        # SEC-01: Validate workspace access
        user = g.user
        is_super_admin = bool(user.get('is_super_admin'))
        user_ws_ids = [str(w) for w in (user.get('workspace_ids') or [])]
        if workspace_id and not is_super_admin and workspace_id not in user_ws_ids:
            return jsonify({'error': 'Acesso negado a este workspace'}), 403

        if workspace_id:
            ws_resp = requests.get(
                f'{_SUPABASE_URL}/rest/v1/workspaces?id=eq.{quote(workspace_id)}&select=name',
                headers=_supabase_headers(),
                timeout=10,
            )
            if ws_resp.ok and ws_resp.json():
                workspace_name = (ws_resp.json()[0].get('name') or '').strip()

        url = (
            f'{_SUPABASE_URL}/rest/v1/chamados_tickets'
            f'?select=status,priority,problemCategory,problemArea,roomName,assignedTo,createdAt,resolvedAt,feedbackRating'
            f'&createdAt=gte.{quote(from_iso)}&createdAt=lte.{quote(to_iso)}'
        )
        if workspace_id:
            url += f'&workspace_id=eq.{quote(workspace_id)}'
        resp = requests.get(url, headers=_supabase_headers(), timeout=15)
        if not resp.ok:
            return jsonify({'error': 'Erro ao gerar o resumo'}), 502

        report = _aggregate_ticket_reports(resp.json() or [])
        html = _build_weekly_report_html(report, workspace_name)

        to = str(body.get('to') or '').strip() or os.environ.get('REPORT_EMAIL_TO', '').strip()
        ok, err = _send_email_via_resend(to, f'LabHub · Resumo semanal de chamados ({report["total"]})', html)
        if not ok:
            return jsonify({'error': err}), 400 if 'não configurado' in err or 'destinatário' in err else 502

        return jsonify({'ok': True, 'total': report['total'], 'sent_to': to})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/chamados/photos/purge', methods=['POST'])
@require_cron
def chamados_photos_purge():
    """Cron diário: apaga do Cloudinary as fotos de chamados fechados há mais de 2 dias.

    Roda via GitHub Actions (photos-cleanup.yml) com Bearer CRON_SECRET.
    Limpa a coluna photos do ticket e photo_urls dos eventos afetados.
    """
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        _ensure_chamados_schema()
        cutoff = (datetime.now(timezone.utc) - timedelta(days=2)).isoformat()
        resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/chamados_tickets'
            f'?status=eq.fechado&closedAt=lt.{quote(cutoff)}&select=id,workspace_id,photos',
            headers=_supabase_headers(),
            timeout=15,
        )
        if not resp.ok:
            return jsonify({'error': 'Erro ao buscar chamados fechados'}), 502
        tickets = resp.json() or []

        deleted = 0
        cleared_tickets = []
        cleared_events = 0

        for t in tickets:
            ticket_photo = str(t.get('photos') or '').strip()
            if ticket_photo and ticket_photo.startswith('https://res.cloudinary.com/') and _cloudinary_destroy(ticket_photo):
                deleted += 1
                cleared_tickets.append(t.get('id'))
                requests.patch(
                    f'{_SUPABASE_URL}/rest/v1/chamados_tickets?id=eq.{quote(t.get("id"))}',
                    headers={**_supabase_headers(), 'Prefer': 'return=minimal'},
                    json={'photos': '', 'updatedAt': datetime.now(timezone.utc).isoformat()},
                    timeout=10,
                )

        ticket_ids = [t.get('id') for t in tickets if t.get('id')]
        if ticket_ids:
            events_resp = requests.get(
                f'{_SUPABASE_URL}/rest/v1/ticket_events'
                f'?ticket_id=in.({",".join(quote(i) for i in ticket_ids)})&select=id,photo_urls',
                headers=_supabase_headers(),
                timeout=15,
            )
            if events_resp.ok:
                for ev in (events_resp.json() or []):
                    try:
                        urls = json.loads(ev.get('photo_urls') or '[]')
                    except (TypeError, ValueError):
                        urls = []
                    cloud_urls = [u for u in urls if isinstance(u, str) and u.startswith('https://res.cloudinary.com/')]
                    if not cloud_urls:
                        continue
                    removed = 0
                    for u in cloud_urls:
                        if _cloudinary_destroy(u):
                            removed += 1
                    if removed:
                        deleted += removed
                        cleared_events += 1
                        remaining = [u for u in urls if u not in cloud_urls]
                        requests.patch(
                            f'{_SUPABASE_URL}/rest/v1/ticket_events?id=eq.{quote(ev.get("id"))}',
                            headers={**_supabase_headers(), 'Prefer': 'return=minimal'},
                            json={'photo_urls': json.dumps(remaining, ensure_ascii=False)},
                            timeout=10,
                        )

        return jsonify({
            'tickets_scanned': len(tickets),
            'tickets_cleared': len(cleared_tickets),
            'events_cleared': cleared_events,
            'photos_deleted': deleted,
        })

    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


# ──────────────────────────────────────────────────────────────────────────────
# Workspace Backups & Audit Logs — backend-only (super_admin)
# ──────────────────────────────────────────────────────────────────────────────

_BACKUP_TTL_DAYS = 2
_UUID_RE = re.compile(
    r'^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$'
)


def _require_super_admin():
    """Return None if the current user is super_admin, otherwise a 403 response."""
    user = getattr(g, 'user', None) or {}
    if not user.get('is_super_admin'):
        return _forbidden('Super admin access required')
    return None


@app.route('/api/admin/backups', methods=['GET'])
@require_auth
def admin_backups_list():
    """List non-expired workspace backups. Super_admin only."""
    admin_err = _require_super_admin()
    if admin_err:
        return admin_err
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/workspace_backups'
            f'?expires_at=gt.{datetime.now(timezone.utc).isoformat()}'
            f'&select=*&order=created_at.desc',
            headers=_supabase_headers(),
            timeout=15,
        )
        if not resp.ok:
            return jsonify({'error': 'Erro ao buscar backups'}), 502
        return jsonify({'backups': resp.json() or []})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/admin/backups/prune', methods=['POST'])
@require_auth
@require_action_rbac('admin.backup.delete', scope='global')
def admin_backups_prune():
    """Delete expired backups. Super_admin only."""
    admin_err = _require_super_admin()
    if admin_err:
        return admin_err
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        resp = requests.delete(
            f'{_SUPABASE_URL}/rest/v1/workspace_backups'
            f'?expires_at=lt.{datetime.now(timezone.utc).isoformat()}',
            headers=_supabase_headers(),
            timeout=15,
        )
        if not resp.ok:
            return jsonify({'error': 'Erro ao limpar backups'}), 502
        return jsonify({'ok': True})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/admin/backups/<backup_id>/restore', methods=['POST'])
@require_auth
@require_action_rbac('admin.backup.restore', scope='global')
def admin_backups_restore(backup_id):
    """Restore a workspace from backup. Super_admin only.

    Flow: validate → read backup → check TTL → upsert workspace →
          audit log → delete consumed backup → return success.
    """
    admin_err = _require_super_admin()
    if admin_err:
        return admin_err
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    if not backup_id or not _UUID_RE.match(backup_id):
        return jsonify({'error': 'ID de backup inválido'}), 400
    try:
        # 1. Read backup
        resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/workspace_backups'
            f'?id=eq.{quote(backup_id)}&select=*',
            headers=_supabase_headers(),
            timeout=15,
        )
        if not resp.ok:
            return jsonify({'error': 'Erro ao buscar backup'}), 502
        rows = resp.json() or []
        if not rows:
            return jsonify({'error': 'Backup não encontrado'}), 404

        backup = rows[0]
        expires_at = backup.get('expires_at', '')
        if expires_at and datetime.fromisoformat(expires_at.replace('Z', '+00:00')) < datetime.now(timezone.utc):
            return jsonify({'error': 'Backup expirado'}), 410

        workspace_data = backup.get('workspace_data')
        if not workspace_data or not isinstance(workspace_data, dict):
            return jsonify({'error': 'Dados do backup inválidos'}), 422

        # 2. Restore workspace via upsert
        upsert_resp = requests.post(
            f'{_SUPABASE_URL}/rest/v1/workspaces',
            headers={**_supabase_headers(), 'Prefer': 'resolution=merge-duplicates,return=minimal'},
            json=workspace_data,
            timeout=15,
        )
        if not upsert_resp.ok:
            return jsonify({'error': 'Falha ao restaurar workspace'}), 502

        # 3. Audit log
        user = g.user or {}
        actor_name = user.get('name') or user.get('id') or 'desconhecido'
        requests.post(
            f'{_SUPABASE_URL}/rest/v1/workspace_audit_logs',
            headers={**_supabase_headers(), 'Prefer': 'return=minimal'},
            json={
                'action': 'restore',
                'workspace_id': workspace_data.get('id'),
                'workspace_name': workspace_data.get('name'),
                'actor_id': g.user_id,
                'actor_name': actor_name,
            },
            timeout=15,
        )

        # 4. Delete consumed backup
        requests.delete(
            f'{_SUPABASE_URL}/rest/v1/workspace_backups?id=eq.{quote(backup_id)}',
            headers={**_supabase_headers(), 'Prefer': 'return=minimal'},
            timeout=15,
        )

        return jsonify({'ok': True, 'workspace_id': workspace_data.get('id')})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/admin/backups/<backup_id>', methods=['DELETE'])
@require_auth
@require_action_rbac('admin.backup.delete', scope='global')
def admin_backups_delete(backup_id):
    """Delete a specific backup. Super_admin only."""
    admin_err = _require_super_admin()
    if admin_err:
        return admin_err
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    if not backup_id or not _UUID_RE.match(backup_id):
        return jsonify({'error': 'ID de backup inválido'}), 400
    try:
        resp = requests.delete(
            f'{_SUPABASE_URL}/rest/v1/workspace_backups?id=eq.{quote(backup_id)}',
            headers={**_supabase_headers(), 'Prefer': 'return=minimal'},
            timeout=15,
        )
        if not resp.ok:
            return jsonify({'error': 'Erro ao excluir backup'}), 502
        return jsonify({'ok': True})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/admin/audit-logs', methods=['GET'])
@require_auth
@require_action_rbac('admin.audit.view', scope='global')
def admin_audit_logs_list():
    """List workspace audit logs (last 100). Super_admin only."""
    admin_err = _require_super_admin()
    if admin_err:
        return admin_err
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    try:
        resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/workspace_audit_logs'
            f'?select=*&order=created_at.desc&limit=100',
            headers=_supabase_headers(),
            timeout=15,
        )
        if not resp.ok:
            return jsonify({'error': 'Erro ao buscar logs'}), 502
        return jsonify({'logs': resp.json() or []})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


@app.route('/api/admin/workspaces/<workspace_id>/delete', methods=['POST'])
@require_auth
@require_action_rbac('admin.workspace.delete', scope='global')
def admin_workspace_delete_with_backup(workspace_id):
    """Delete a workspace with backup + audit. Super_admin only.

    Atomic flow: validate → backup → delete workspace → audit → prune.
    """
    admin_err = _require_super_admin()
    if admin_err:
        return admin_err
    if not _require_supabase():
        return jsonify({'error': 'Supabase não configurado'}), 503
    if not workspace_id or not _UUID_RE.match(workspace_id):
        return jsonify({'error': 'ID de workspace inválido'}), 400

    try:
        # 1. Fetch workspace data for backup
        ws_resp = requests.get(
            f'{_SUPABASE_URL}/rest/v1/workspaces?id=eq.{quote(workspace_id)}&select=*',
            headers=_supabase_headers(),
            timeout=15,
        )
        if not ws_resp.ok:
            return jsonify({'error': 'Erro ao buscar workspace'}), 502
        ws_rows = ws_resp.json() or []
        if not ws_rows:
            return jsonify({'error': 'Workspace não encontrado'}), 404

        ws_data = ws_rows[0]
        ws_name = ws_data.get('name', '')

        # 2. Create backup
        expires_at = (datetime.now(timezone.utc) + timedelta(days=_BACKUP_TTL_DAYS)).isoformat()
        user = g.user or {}
        actor_name = user.get('name') or user.get('id') or 'desconhecido'
        backup_resp = requests.post(
            f'{_SUPABASE_URL}/rest/v1/workspace_backups',
            headers={**_supabase_headers(), 'Prefer': 'return=minimal'},
            json={
                'workspace_id': workspace_id,
                'workspace_name': ws_name,
                'workspace_data': ws_data,
                'deleted_by': g.user_id,
                'deleted_by_name': actor_name,
                'expires_at': expires_at,
            },
            timeout=15,
        )
        if not backup_resp.ok:
            return jsonify({'error': 'Falha ao criar backup'}), 502

        # 3. Delete workspace
        del_resp = requests.delete(
            f'{_SUPABASE_URL}/rest/v1/workspaces?id=eq.{quote(workspace_id)}',
            headers={**_supabase_headers(), 'Prefer': 'return=minimal'},
            timeout=15,
        )
        if not del_resp.ok:
            return jsonify({'error': 'Falha ao excluir workspace'}), 502

        # 4. Audit log
        requests.post(
            f'{_SUPABASE_URL}/rest/v1/workspace_audit_logs',
            headers={**_supabase_headers(), 'Prefer': 'return=minimal'},
            json={
                'action': 'delete',
                'workspace_id': workspace_id,
                'workspace_name': ws_name,
                'actor_id': g.user_id,
                'actor_name': actor_name,
            },
            timeout=15,
        )

        # 5. Prune expired (best-effort)
        requests.delete(
            f'{_SUPABASE_URL}/rest/v1/workspace_backups'
            f'?expires_at=lt.{datetime.now(timezone.utc).isoformat()}',
            headers={**_supabase_headers(), 'Prefer': 'return=minimal'},
            timeout=15,
        )

        return jsonify({'ok': True, 'backup_name': ws_name})
    except Exception as e:
        logger.error("Erro interno na API: %s", e)
        return jsonify({'error': 'Erro interno'}), 500


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000, use_reloader=False)

