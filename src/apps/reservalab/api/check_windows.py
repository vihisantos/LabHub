#!/usr/bin/env python
"""Verificar onde estão as informações do Windows"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from api.invent import INVENTARIO_URL
import requests
from io import BytesIO
from openpyxl import load_workbook

response = requests.get(INVENTARIO_URL, timeout=30)
wb = load_workbook(BytesIO(response.content), data_only=True)
ws = wb['Consultórios CIS med']

print("Verificando colunas para informações do Windows:\n")
print("Headers (row 1):")
for cell in ws[1]:
    if cell.value:
        print(f"  Col {cell.column}: {cell.value}")

print("\nPrimeiras 3 linhas - verificando todas as colunas:")
for i in range(2, 5):
    row = [cell.value for cell in ws[i]]
    print(f"\nLinha {i}:")
    for j, val in enumerate(row):
        if val:
            print(f"  Col {j+1}: {val}")
