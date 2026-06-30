from openpyxl import load_workbook
import os
from datetime import date, timedelta

ARQUIVO_URL = r'C:\Users\manoel.santana\OneDrive - Administrativo - Ânima Educação\Área de Trabalho\planilha\Reservas Labs - 2026.xlsx'

wb = load_workbook(ARQUIVO_URL, data_only=True)
ws = wb['RESERVA LAB. INFORMÁTICA']

hoje = date.today()
print(f"Hoje: {hoje}")

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=5, values_only=True)):
    print(f"\n--- Row {i+2} ---")
    print(f"0(A): {row[0]}")
    print(f"1(B): {row[1]}")
    print(f"2(C): {row[2]}")
    print(f"3(D): {row[3]} - type: {type(row[3])}")
    print(f"4(E): {row[4]}")
    print(f"5(F): {row[5]}")
    print(f"6(G): {row[6]}")
    print(f"8(I): {row[8]}")
    
    data_reserva = row[3]
    if data_reserva:
        if hasattr(data_reserva, 'date'):
            data = data_reserva.date()
            print(f"Data convertida: {data}, igual hoje? {data == hoje}")