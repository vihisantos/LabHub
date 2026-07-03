import time
import json
from flask import Flask, jsonify, request
from openpyxl import load_workbook
from datetime import date, timedelta, datetime
from flask_cors import CORS
import os
import sys
import logging
import requests
import re
import hashlib
from zoneinfo import ZoneInfo
from urllib.parse import quote
from io import BytesIO
from dotenv import load_dotenv
from upstash_redis import Redis
from pywebpush import webpush

# Detecta caminho correto quando rodando de exe
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Carrega .env do diretório do exe
env_path = os.path.join(BASE_DIR, '.env')
if os.path.exists(env_path):
    load_dotenv(env_path)
else:
    load_dotenv()

def get_now_sp():
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo('America/Sao_Paulo'))
    except Exception:
        from datetime import timezone
        return datetime.now(timezone(timedelta(hours=-3)))

def get_today_sp():
    return get_now_sp().date()


class JSONFormatter(logging.Formatter):
    def format(self, record):
        return json.dumps({
            "timestamp": self.formatTime(record),
            "level": record.levelname,
            "message": record.getMessage(),
            "module": record.module,
            "function": record.funcName,
            "line": record.lineno
        })

logging.basicConfig(
    level=logging.INFO,
    format='%(message)s',
    stream=sys.stdout
)
logger = logging.getLogger(__name__)
for h in logger.handlers:
    h.setFormatter(JSONFormatter())

CACHE_TTL = 60  # segundos
if os.environ.get('VERCEL'):
    CACHE_FILE = '/tmp/.cache_reservas.json'
else:
    CACHE_FILE = os.path.join(BASE_DIR, '.cache_reservas.json')

def get_cached_reservas():
    try:
        if os.path.exists(CACHE_FILE):
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                cache = json.load(f)
            if time.time() - cache.get('timestamp', 0) < CACHE_TTL:
                logger.info("Usando cache")
                return cache.get('data')
    except Exception as e:
        logger.error(f"Erro ao ler cache: {e}")
    return None

class DateEncoder(json.JSONEncoder):
    """Serializa objetos date/datetime para string ISO automaticamente."""
    def default(self, obj):
        if isinstance(obj, (date, datetime)):
            return obj.strftime('%d/%m/%Y')
        return super().default(obj)

def set_cached_reservas(data):
    try:
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump({'data': data, 'timestamp': time.time()}, f, cls=DateEncoder)
    except Exception as e:
        logger.error(f"Erro ao salvar cache: {e}")

app = Flask(__name__)
CORS(app)

VAPID_PUBLIC_KEY = "BPySgynYSvDIXSa3haOi3GyDolJqGhMyCdWdZfurxZA-OFySm-rZaIWXDJss2sEV2ngRe2Zp6_1gvxizQ9v2s8g"
VAPID_PRIVATE_KEY = "8Cfm5ZvDQzP5DP2oyOzyZ-K2ks06K_OQoSlKcpCkWaI"
VAPID_CLAIMS = {"sub": "mailto:admin@reservaslab.com"}

_upstash_url = os.environ.get('UPSTASH_REDIS_REST_URL')
_upstash_token = os.environ.get('UPSTASH_REDIS_REST_TOKEN')
redis = Redis(url=_upstash_url, token=_upstash_token) if _upstash_url and _upstash_token else None

ARQUIVO_URL = os.environ.get('SHAREPOINT_URL', '')
if not ARQUIVO_URL:
    logger.error("URL da planilha nao configurada no .env!")
    ARQUIVO_URL = None

def get_reservas():
    cached = get_cached_reservas()
    if cached:
        return cached
    
    reservas_hoje = []
    reservas_semana = []
    
    if ARQUIVO_URL:
        try:
            logger.info("Baixando planilha do SharePoint...")
            response = requests.get(ARQUIVO_URL, timeout=30)
            response.raise_for_status()
            
            try:
                wb = load_workbook(BytesIO(response.content))
                logger.info("Planilha carregada!")
            except Exception as e:
                logger.error(f"Erro ao carregar planilha: {e}")
                wb = None
            
            if wb:
                try:
                    ws = wb['RESERVA LAB. INFORMÁTICA']
                except KeyError:
                    logger.error("Aba 'RESERVA LAB. INFORMÁTICA' não encontrada")
                    ws = None
                
                if ws:
                    hoje = get_today_sp()
                    fim_semana = hoje + timedelta(days=7)
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        try:
                            reserva_feita_por = row[0]
                            professor_resp = row[1]
                            email = row[2]
                            data_reserva = row[3]
                            horario = row[4]
                            alunos = row[5]
                            obs = row[6]
                            lab = row[8]
                            
                            if data_reserva is None:
                                continue
                            
                            data = None
                            if hasattr(data_reserva, 'date'):
                                try:
                                    data = data_reserva.date()
                                except:
                                    pass
                            
                            if data is None and isinstance(data_reserva, str):
                                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                                    try:
                                        data = datetime.strptime(data_reserva, fmt).date()
                                        break
                                    except:
                                        continue
                            
                            if data is None:
                                continue
                            
                            lab_normalizado = re.sub(r'\s+', ' ', str(lab)).strip() if lab else ''
                            lab_list = []
                            
                            if re.search(r'lab\s*0?\s*1', lab_normalizado, re.IGNORECASE):
                                lab_list.append('LAB01')
                            if re.search(r'lab\s*0?\s*2', lab_normalizado, re.IGNORECASE):
                                lab_list.append('LAB02')
                            
                            reserva = {
                                'responsavel': professor_resp,
                                'email': email,
                                'horario': horario,
                                'alunos': alunos,
                                'observacao': obs,
                                'lab': lab,
                                'labs': lab_list,
                                'data': data,
                                'reserva_feita_por': reserva_feita_por,
                                'origem': 'planilha'
                            }
                            
                            if data == hoje:
                                reservas_hoje.append(reserva)
                            elif hoje < data <= fim_semana:
                                reservas_semana.append(reserva)
                        except Exception as e:
                            logger.warning(f"Erro processando linha: {e}")
                            continue
        except Exception as e:
            logger.error(f"Erro geral ao processar planilha: {e}")
    
    logger.info(f"Planilha: {len(reservas_hoje)} hoje, {len(reservas_semana)} semana")
    
    result = (reservas_hoje, reservas_semana)
    set_cached_reservas(result)
    return result

@app.route('/api/reservas', methods=['GET'])
def api_reservas():
    try:
        reservas_hoje, reservas_semana = get_reservas()
        
        for r in reservas_hoje:
            if isinstance(r.get('data'), (date, datetime)):
                r['data'] = r['data'].strftime('%d/%m/%Y')
        
        for r in reservas_semana:
            if isinstance(r.get('data'), (date, datetime)):
                r['data'] = r['data'].strftime('%d/%m/%Y')

        lab1 = [r for r in reservas_hoje if 'LAB01' in r.get('labs', [])]
        lab2 = [r for r in reservas_hoje if 'LAB02' in r.get('labs', [])]
        
        meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
                 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        hoje = get_today_sp()
        data_formatada = f"{hoje.day:02d} de {meses[hoje.month-1]} de {hoje.year}"
        
        cache_info = {}
        if os.path.exists(CACHE_FILE):
            try:
                with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                    cache_data = json.load(f)
                    cache_info = {'timestamp': cache_data.get('timestamp', 0)}
            except:
                pass

        return jsonify({
            'lab1_reservas': lab1,
            'lab2_reservas': lab2,
            'reservas_semana': reservas_semana,
            'data': data_formatada,
            'cache_info': cache_info
        })
    except Exception as e:
        logger.error(f"Erro na rota api_reservas: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/health')
def health():
    cache_data = None
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)
        except Exception:
            pass
    return jsonify({
        'status': 'ok',
        'cache': {
            'ativo': True,
            'ttl': CACHE_TTL,
            'timestamp': cache_data.get('timestamp', 0) if cache_data else 0
        },
        'url_configurada': bool(ARQUIVO_URL)
    })

# ─── Push Notifications ──────────────────────────────────────────

def parse_horario_inicio(horario):
    if not horario:
        return None
    s = str(horario).strip().lower()
    # Pega apenas a primeira parte do horário (antes de "até", "as", "às", "a", "-", etc.)
    parts = re.split(r'\s+(?:até|ate|as|às|a)\s+|(?:\s+-\s+)|-', s)
    s = parts[0].strip() if parts else s
    
    # Formatos: "11:30", "11h30", "11h", "11", "09:35h"
    match = re.match(r'^(\d{1,2})(?:\s*[:h]\s*(\d{2}))?', s)
    if match:
        hora = int(match.group(1))
        minuto = int(match.group(2)) if match.group(2) else 0
        return hora * 60 + minuto
    return None

@app.route('/api/push/subscribe', methods=['POST'])
def push_subscribe():
    if not redis:
        return jsonify({'error': 'Redis not configured'}), 500
    try:
        sub = request.get_json()
        sub_key = hashlib.sha256(json.dumps(sub, sort_keys=True).encode()).hexdigest()
        redis.sadd('push:subscribers', json.dumps({'key': sub_key, **sub}))
        logger.info(f"Push subscriber added: {sub_key[:8]}...")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Push subscribe error: {e}")
        return jsonify({'error': str(e)}), 500

def push_notify(sub, title, body, url='/'):
    try:
        webpush(
            subscription_info=sub,
            data=json.dumps({'title': title, 'body': body, 'url': url}),
            vapid_private_key=VAPID_PRIVATE_KEY,
            vapid_claims=VAPID_CLAIMS,
            ttl=86400
        )
        return True
    except Exception as e:
        logger.warning(f"Push send error: {e}")
        return False

@app.route('/api/push/test', methods=['GET'])
def push_test():
    if not redis:
        return jsonify({'error': 'Redis not configured'}), 500
    subs = redis.smembers('push:subscribers')
    if not subs:
        return jsonify({'message': 'No subscribers'})
    count = 0
    for raw in subs:
        sub = json.loads(raw) if isinstance(raw, str) else raw
        ok = push_notify(sub, 'Teste Reservas Lab', 'Notificação push funcionando! 🔔')
        if ok:
            count += 1
    return jsonify({'sent': count, 'total': len(subs)})


@app.route('/api/push/check', methods=['GET'])
def push_check():
    if not redis:
        return jsonify({'error': 'Redis not configured'}), 500
    try:
        today = get_today_sp()
        reservas_hoje, _ = get_reservas()
        
        agora = get_now_sp()
        agora_min = agora.hour * 60 + agora.minute
        limite_min = agora_min + 15
        
        subs_raw = redis.smembers('push:subscribers')
        if not subs_raw:
            return jsonify({'message': 'No subscribers', 'sent': 0})
        
        subs = [json.loads(s) if isinstance(s, str) else s for s in subs_raw]
        sent = 0
        
        for r in reservas_hoje:
            data_reserva = r.get('data')
            if isinstance(data_reserva, str):
                try:
                    data_reserva = datetime.strptime(data_reserva, '%d/%m/%Y').date()
                except:
                    continue
            if data_reserva != today:
                continue
            
            inicio = parse_horario_inicio(r.get('horario', ''))
            if inicio is None:
                continue
            
            if agora_min <= inicio <= limite_min:
                notify_id = hashlib.md5(f"{r['lab']}|{r['horario']}|{r.get('responsavel','')}".encode()).hexdigest()
                
                already = redis.get(f'push:sent:{notify_id}')
                if already:
                    continue
                
                title = f"{' '.join(r.get('labs', [r.get('lab', 'Lab')]))}"
                alunos = r.get('alunos', '')
                evento = r.get('observacao', '').strip()
                body = f"{r['horario']} — {r.get('responsavel', 'Professor') or 'Sem professor'}"
                if evento:
                    body += f" — {evento}"
                if alunos:
                    body += f" — {alunos} alunos"
                
                for sub in subs:
                    push_notify(sub, title, body)
                
                redis.setex(f'push:sent:{notify_id}', 7200, '1')
                sent += 1
                logger.info(f"Push sent for {title} at {r['horario']}")
        
        # ─── Push para reservas de tablets ──────────────────────────
        supabase_url = os.environ.get('SUPABASE_URL', '')
        supabase_key = os.environ.get('SUPABASE_SERVICE_KEY', '')
        if supabase_url and supabase_key:
            try:
                today_str = today.strftime('%Y-%m-%d')
                tomorrow_str = (today + timedelta(days=1)).strftime('%Y-%m-%d')
                tablet_url = (
                    f"{supabase_url}/rest/v1/tablet_reservations"
                    f"?select=*&order=horario_inicio.asc"
                    f"&status=eq.{quote('ativa')}"
                    f"&horario_inicio=gte.{quote(today_str + 'T00:00:00Z')}"
                    f"&horario_inicio=lt.{quote(tomorrow_str + 'T00:00:00Z')}"
                )
                tablet_resp = requests.get(
                    tablet_url,
                    headers={
                        'apikey': supabase_key,
                        'Authorization': f'Bearer {supabase_key}',
                    },
                    timeout=10
                )
                if tablet_resp.ok:
                    tablets_hoje = tablet_resp.json()
                    agora = get_now_sp()
                    agora_min = agora.hour * 60 + agora.minute
                    limite_min = agora_min + 15
                    
                    for t in tablets_hoje:
                        inicio_str = t.get('horario_inicio', '')
                        if not inicio_str:
                            continue
                        try:
                            dt = datetime.fromisoformat(inicio_str.replace('Z', '+00:00'))
                            dt_sp = dt.astimezone(ZoneInfo('America/Sao_Paulo'))
                            inicio_min = dt_sp.hour * 60 + dt_sp.minute
                        except:
                            continue
                        
                        if agora_min <= inicio_min <= limite_min:
                            notify_id = hashlib.md5(f"tablet|{t['id']}".encode()).hexdigest()
                            already = redis.get(f'push:sent:{notify_id}')
                            if already:
                                continue
                            
                            title = f"Tablets — {t.get('sala', 'Sala')}"
                            evento = t.get('finalidade', '').strip()
                            body = f"{t.get('professor', 'Professor')}"
                            if evento:
                                body += f" — {evento}"
                            body += f" — {t.get('quantidade_tablets', '?')} tablets"
                            
                            for sub in subs:
                                push_notify(sub, title, body)
                            
                            redis.setex(f'push:sent:{notify_id}', 7200, '1')
                            sent += 1
                            logger.info(f"Push sent for tablet: {title} at {inicio_str}")
            except Exception as e:
                logger.error(f"Tablet push check error: {e}")
        
        return jsonify({'checked': True, 'sent': sent, 'subscribers': len(subs)})
    except Exception as e:
        logger.error(f"Push check error: {e}")
        return jsonify({'error': str(e)}), 500

# ─── Notificações de Empréstimos (Stock) ─────────────────────────

def _get_subs():
    raw = redis.smembers('push:subscribers') if redis else []
    return [json.loads(s) if isinstance(s, str) else s for s in raw]


@app.route('/api/push/notify-loan', methods=['POST'])
def push_notify_loan():
    if not redis:
        return jsonify({'error': 'Redis not configured'}), 500
    try:
        body = request.get_json()
        item_name = body.get('itemName', 'Item')
        borrowed_by = body.get('borrowedBy', 'Alguém')
        expected_return = body.get('expectedReturnAt', '')

        title = f"📦 Empréstimo: {item_name}"
        msg = f"Emprestado para {borrowed_by}"
        if expected_return:
            msg += f" — Devolução até {expected_return[:10]}"

        subs = _get_subs()
        for sub in subs:
            push_notify(sub, title, msg)

        logger.info(f"Loan notify: {title}")
        return jsonify({'sent': len(subs)})
    except Exception as e:
        logger.error(f"notify-loan error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/push/notify-return', methods=['POST'])
def push_notify_return():
    if not redis:
        return jsonify({'error': 'Redis not configured'}), 500
    try:
        body = request.get_json()
        item_name = body.get('itemName', 'Item')
        returned_by = body.get('returnedBy', 'Alguém')

        title = f"✅ Devolução: {item_name}"
        msg = f"Devolvido por {returned_by}"

        subs = _get_subs()
        for sub in subs:
            push_notify(sub, title, msg)

        logger.info(f"Return notify: {title}")
        return jsonify({'sent': len(subs)})
    except Exception as e:
        logger.error(f"notify-return error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/push/check-overdue', methods=['GET'])
def push_check_overdue():
    if not redis:
        return jsonify({'error': 'Redis not configured'}), 500
    try:
        supabase_url = os.environ.get('SUPABASE_URL', '')
        supabase_key = os.environ.get('SUPABASE_SERVICE_KEY', '')
        if not supabase_url or not supabase_key:
            return jsonify({'error': 'Supabase not configured'}), 500

        agora = get_now_sp()
        limite_12h = agora + timedelta(hours=12)

        headers = {
            'apikey': supabase_key,
            'Authorization': f'Bearer {supabase_key}',
        }

        url = (
            f"{supabase_url}/rest/v1/stock_movements"
            f"?select=*"
            f"&type=eq.{quote('emprestimo')}"
            f"&returnedAt=is.null"
        )
        resp = requests.get(url, headers=headers, timeout=10)
        if not resp.ok:
            logger.error(f"check-overdue Supabase error: {resp.status_code}")
            return jsonify({'error': 'Supabase query failed'}), 500

        all_loans = resp.json()
        subs = _get_subs()
        sent = 0
        found = 0

        for loan in all_loans:
            expected_raw = loan.get('expectedReturnAt')
            if not expected_raw:
                continue
            try:
                if 'T' in expected_raw:
                    dt = datetime.fromisoformat(expected_raw.replace('Z', '+00:00'))
                else:
                    dt = datetime.strptime(expected_raw[:10], '%Y-%m-%d').replace(hour=23, minute=59)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=ZoneInfo('America/Sao_Paulo'))
                else:
                    dt = dt.astimezone(ZoneInfo('America/Sao_Paulo'))
            except Exception:
                continue

            if not (agora <= dt <= limite_12h):
                continue

            found += 1
            nid = hashlib.md5(f"overdue|{loan['id']}".encode()).hexdigest()
            if redis.get(f'push:sent:{nid}'):
                continue

            item_name = loan.get('itemName', 'Item')
            borrowed_by = loan.get('borrowedBy', 'Alguém')

            title = f"⏰ Prazo de devolução: {item_name}"
            msg = f"Emprestado para {borrowed_by} — Vence {expected_raw[:10]}"
            for sub in subs:
                push_notify(sub, title, msg)

            redis.setex(f'push:sent:{nid}', 43200, '1')
            sent += 1
            logger.info(f"Overdue notify: {item_name}")

        return jsonify({'checked': True, 'sent': sent, 'found': found, 'subscribers': len(subs)})
    except Exception as e:
        logger.error(f"check-overdue error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/push/check-pcare', methods=['GET'])
def push_check_pcare():
    if not redis:
        return jsonify({'error': 'Redis not configured'}), 500
    try:
        supabase_url = os.environ.get('SUPABASE_URL', '')
        supabase_key = os.environ.get('SUPABASE_SERVICE_KEY', '')
        if not supabase_url or not supabase_key:
            return jsonify({'error': 'Supabase not configured'}), 500

        agora = get_now_sp()
        hoje_str = agora.strftime('%Y-%m-%d')
        amanha_str = (agora + timedelta(days=1)).strftime('%Y-%m-%d')

        base_headers = {
            'apikey': supabase_key,
            'Authorization': f'Bearer {supabase_key}',
        }
        subs = _get_subs()
        sent = 0

        # ── Estoque baixo de peças ──
        parts_headers = base_headers
        try:
            pr = requests.get(
                f"{supabase_url}/rest/v1/parts?select=*",
                headers=parts_headers, timeout=10
            )
            if pr.ok:
                for part in pr.json():
                    qty = part.get('quantity', 0)
                    min_qty = part.get('minQuantity', 0)
                    if qty < min_qty:
                        nid = hashlib.md5(f"pcare|part|{part['id']}".encode()).hexdigest()
                        if redis.get(f'push:sent:{nid}'):
                            continue
                        title = f"🔧 Estoque baixo: {part.get('name', 'Peça')}"
                        msg = f"Quantidade: {qty} | Mínimo: {min_qty}"
                        for sub in subs:
                            push_notify(sub, title, msg)
                        redis.setex(f'push:sent:{nid}', 86400, '1')
                        sent += 1
                        logger.info(f"Low stock: {part.get('name')}")
        except Exception as e:
            logger.error(f"check-pcare parts error: {e}")

        # ── Manutenções agendadas ──
        try:
            mr = requests.get(
                f"{supabase_url}/rest/v1/maintenance"
                f"?select=*"
                f"&completed=eq.false"
                f"&scheduledDate=gte.{quote(hoje_str)}"
                f"&scheduledDate=lte.{quote(amanha_str)}",
                headers=parts_headers, timeout=10
            )
            if mr.ok:
                for m in mr.json():
                    nid = hashlib.md5(f"pcare|maint|{m['id']}".encode()).hexdigest()
                    if redis.get(f'push:sent:{nid}'):
                        continue
                    title = f"🔧 Manutenção: {m.get('pcNumber', 'PC')}"
                    msg = f"{m.get('labName', 'Lab')} — {m.get('type', '')} — {m.get('scheduledDate', '')[:10]}"
                    for sub in subs:
                        push_notify(sub, title, msg)
                    redis.setex(f'push:sent:{nid}', 86400, '1')
                    sent += 1
                    logger.info(f"Maintenance: {m.get('pcNumber')}")
        except Exception as e:
            logger.error(f"check-pcare maintenance error: {e}")

        return jsonify({'checked': True, 'sent': sent, 'subscribers': len(subs)})
    except Exception as e:
        logger.error(f"check-pcare error: {e}")
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000, use_reloader=False)
