"""
Módulo de lógica do Inventário.
Apenas funções puras — sem Flask aqui.
Importado pelo app.py principal.
"""

import time
import json
import os
import logging
import requests
import re
from io import BytesIO
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

logger = logging.getLogger(__name__)

INVENTARIO_URL = os.environ.get('INVENTARIO_URL', '')

CACHE_TTL = 1800  # 30 minutos

if os.environ.get('VERCEL'):
    CACHE_FILE = '/tmp/.cache_inventario.json'
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    CACHE_FILE = os.path.join(BASE_DIR, '.cache_inventario.json')


def get_cached_inventario(aba=None):
    try:
        if os.path.exists(CACHE_FILE):
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                cache = json.load(f)
            if time.time() - cache.get('timestamp', 0) < CACHE_TTL:
                data = cache.get('data')
                # Se pediu uma aba específica, verifica se o cache bate
                if aba and data and data.get('aba') == aba:
                    logger.info("Usando cache do inventário (aba específica)")
                    return data
                elif not aba and data:
                    logger.info("Usando cache do inventário (todas as abas)")
                    return data
    except Exception as e:
        logger.error(f"Erro ao ler cache do inventário: {e}")
    return None


class DateEncoder(json.JSONEncoder):
    """Serializa objetos date/datetime para string ISO automaticamente."""
    def default(self, obj):
        from datetime import date, datetime
        if isinstance(obj, (date, datetime)):
            return obj.strftime('%d/%m/%Y')
        return super().default(obj)

def set_cached_inventario(data):
    try:
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump({'data': data, 'timestamp': time.time()}, f, ensure_ascii=False, cls=DateEncoder)
    except Exception as e:
        logger.error(f"Erro ao salvar cache do inventário: {e}")


def extrair_numero(valor):
    if valor is None:
        return 0
    try:
        return int(str(valor).strip())
    except Exception:
        nums = re.findall(r'\d+', str(valor))
        return int(nums[0]) if nums else 0

def _detectar_tipo_aba(ws):
    """Detecta o tipo de aba baseado nos cabeçalhos."""
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[cell.column] = str(cell.value).strip().lower()
    
    header_text = ' '.join(headers.values())
    
    if 'processador' in header_text or 'service tag' in header_text:
        return 'consultorios'
    elif 'localizacao' in header_text or 'tipo de dispositivo' in header_text:
        return 'lab_informatica'
    elif 'ambiente' in header_text and any('consult' in h or 'cis' in h for h in headers.values()):
        return 'consultorios'  # CIS Med tem coluna AMBIENTE
    elif 'sala' in header_text and 'ambiente' in header_text:
        return 'pcs_sala'
    else:
        return 'desconhecido'

def _mapear_item_consultorios(row, i):
    """Mapeia uma linha da aba CIS med."""
    # 0: Processador, 1: Service TAG, 2: Hostname, 3: Local, 4: Ambiente, 
    # 5: ID cadeado Seg., 6: MAC, 7: Armazenamento, 8: Tipo de Disco, 
    # 9: Memoria Ram, 10: Sistema Operacional, 11: Observações
    
    processador = str(row[0]).strip() if len(row) > 0 and row[0] else ''
    service_tag = str(row[1]).strip() if len(row) > 1 and row[1] else ''
    hostname = str(row[2]).strip() if len(row) > 2 and row[2] else ''
    local = str(row[3]).strip() if len(row) > 3 and row[3] else ''
    ambiente = str(row[4]).strip() if len(row) > 4 and row[4] else ''
    
    mac = str(row[6]).strip() if len(row) > 6 and row[6] else ''
    armazenamento = str(row[7]).strip() if len(row) > 7 and row[7] else ''
    tipo_disco = str(row[8]).strip() if len(row) > 8 and row[8] else ''
    memoria_ram = str(row[9]).strip() if len(row) > 9 and row[9] else ''
    sistema_op = str(row[10]).strip() if len(row) > 10 and row[10] else ''
    observacoes = str(row[11]).strip() if len(row) > 11 and row[11] else ''
    
    nome = f"Consultório {local}" if local else (hostname or service_tag or processador)
    categoria = 'Computador' if any(
        x in processador.lower() for x in ['i5', 'i7', 'dell', 'lenovo', 'hp']
    ) else 'Equipamento'
    
    return {
        'id': service_tag or hostname or processador or str(i),
        'nome': nome,
        'categoria': categoria,
        'quantidade': 1,
        'disponivel': 1,
        'local': local or hostname,
        'observacao': f"{armazenamento} {tipo_disco} {memoria_ram} | {observacoes}".strip(),
        'service_tag': service_tag,
        'hostname': hostname,
        'mac': mac,
        'ip': '',
        'processador': processador,
        'armazenamento': armazenamento,
        'tipo_disco': tipo_disco,
        'memoria_ram': memoria_ram,
        'ambiente': ambiente,
        'sistema_operacional': sistema_op
    }

def _mapear_item_lab_informatica(row, i):
    """Mapeia uma linha das abas laboratorio de Informatica ou Estrutura."""
    # 0: Nome, 1: LOCALIZAÇÃO, 2: AMBIENTE, 3: Patrimonio, 4: TIPO DE AQUISIÇÃO, 
    # 5: TIPO DE DISPOSITIVO, 6: PATRIMÔNIO, 7: NÚMERO DE SÉRIE, 8: STATUS, 
    # 9: MODELO, 10: FABRICANTE, 11: MAC
    
    nome_notebook = str(row[0]).strip() if len(row) > 0 and row[0] else ''
    local = str(row[1]).strip() if len(row) > 1 and row[1] else ''
    tipo_dispositivo = str(row[5]).strip() if len(row) > 5 and row[5] else ''
    patrimonio_num = str(row[6]).strip() if len(row) > 6 and row[6] else ''
    numero_serie = str(row[7]).strip() if len(row) > 7 and row[7] else ''
    status = str(row[8]).strip() if len(row) > 8 and row[8] else ''
    modelo = str(row[9]).strip() if len(row) > 9 and row[9] else ''
    fabricante = str(row[10]).strip() if len(row) > 10 and row[10] else ''
    mac = str(row[11]).strip() if len(row) > 11 and row[11] else ''
    
    nome = nome_notebook if nome_notebook else f"{tipo_dispositivo} - {local}"
    
    return {
        'id': patrimonio_num or numero_serie or nome_notebook or str(i),
        'nome': nome,
        'categoria': tipo_dispositivo or 'Equipamento',
        'quantidade': 1,
        'disponivel': 1 if status.lower() != 'defeito' else 0,
        'local': local,
        'observacao': f"Status: {status} | Modelo: {modelo}" if status else f"Modelo: {modelo}",
        'service_tag': numero_serie,
        'hostname': nome_notebook,
        'mac': mac,
        'ip': '',
        'patrimonio': patrimonio_num,
        'modelo': modelo,
        'ambiente': str(row[2]).strip() if len(row) > 2 else '',
        'status': status,
        'fabricante': fabricante
    }

def _mapear_item_pcs_sala(row, i):
    """Mapeia uma linha da aba PCs sala de aula."""
    # 0: Sala (ex: Sala 101), 1: Ambiente
    sala = str(row[0]).strip() if len(row) > 0 and row[0] else ''
    ambiente = str(row[1]).strip() if len(row) > 1 and row[1] else ''
    
    return {
        'id': f"PC-SALA-{sala}-{i}",
        'nome': f"PC - {sala}",
        'categoria': 'Computador',
        'quantidade': 1,
        'disponivel': 1,
        'local': sala,
        'observacao': f"Ambiente: {ambiente}",
        'service_tag': '',
        'hostname': '',
        'mac': '',
        'ip': '',
        'ambiente': ambiente
    }

def get_inventario_data(aba=None):
    """
    Retorna dados do inventário da planilha Excel hospedada no SharePoint.
    Nunca lança exceção — sempre retorna um dict com 'error' em caso de falha.
    """
    if not INVENTARIO_URL:
        return {
            'error': 'URL do inventário não configurada (INVENTARIO_URL no .env)',
            'itens': [],
            'total': 0,
            'abas_disponiveis': []
        }

    cached = get_cached_inventario(aba)
    if cached:
        return cached

    try:
        logger.info("Baixando planilha de inventário do SharePoint...")
        response = requests.get(INVENTARIO_URL, timeout=30)
        response.raise_for_status()

        wb = load_workbook(BytesIO(response.content), data_only=True)
        abas_disponiveis = wb.sheetnames
        logger.info(f"Planilha carregada. Abas disponíveis: {abas_disponiveis}")

        ws = None
        if aba:
            aba_clean = aba.lower().strip()
            # 1. Tenta match exato ou case-insensitive simples
            abas_map = {a.lower().strip(): a for a in abas_disponiveis}
            if aba_clean in abas_map:
                ws = wb[abas_map[aba_clean]]
            else:
                # 2. Tenta match por substring (ex: "Consultórios CIS med" contem "CIS med")
                for real_name in abas_disponiveis:
                    real_clean = real_name.lower().strip()
                    if real_clean in aba_clean or aba_clean in real_clean:
                        logger.info(f"Match aproximado: '{aba}' -> '{real_name}'")
                        ws = wb[real_name]
                        break
        
        if not ws:
            if aba:
                logger.warning(f"Aba '{aba}' não encontrada. Tentando padrão...")
            # 3. Fallback: procura abas que contenham palavras-chave
            keywords = ['cis', 'consultorio', 'pc', 'informatica', 'laerdal']
            for kw in keywords:
                for real_name in abas_disponiveis:
                    if kw in real_name.lower():
                        ws = wb[real_name]
                        break
                if ws: break
            
            if not ws:
                ws = wb[abas_disponiveis[0]]

        itens = []
        tipo_aba = _detectar_tipo_aba(ws)
        logger.info(f"Tipo de aba detectado: {tipo_aba} para '{ws.title}'")
        
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # Pula linhas completamente vazias
            if not any(cell for cell in row):
                continue
            if not (row[0] or (len(row) > 1 and row[1])):
                continue
            
            if tipo_aba == 'consultorios':
                item = _mapear_item_consultorios(row, i)
            elif tipo_aba == 'lab_informatica':
                item = _mapear_item_lab_informatica(row, i)
            elif tipo_aba == 'pcs_sala':
                item = _mapear_item_pcs_sala(row, i)
            else:
                # Fallback para o mapeamento antigo (compatibilidade)
                consultorio = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                service_tag = str(row[0]).strip() if row[0] else ''
                hostname = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                item = {
                    'id': service_tag or hostname or str(i),
                    'nome': f"Consultório {consultorio}" if consultorio else (hostname or service_tag),
                    'categoria': 'Computador' if any(
                        x in service_tag.lower() for x in ['i5', 'i7', 'dell', 'lenovo', 'hp']
                    ) else 'Equipamento',
                    'quantidade': 1,
                    'disponivel': 0 if (len(row) > 3 and row[3]) else 1,
                    'local': consultorio,
                    'observacao': str(row[6]).strip() if len(row) > 6 and row[6] else '',
                    'service_tag': service_tag,
                    'hostname': hostname,
                    'mac': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                    'ip': str(row[5]).strip() if len(row) > 5 and row[5] else ''
                }
            
            itens.append(item)

        result = {
            'itens': itens,
            'total': len(itens),
            'aba': ws.title,
            'abas_disponiveis': abas_disponiveis
        }

        set_cached_inventario(result)
        logger.info(f"Inventário: {len(itens)} itens carregados da aba '{ws.title}'")
        return result

    except Exception as e:
        logger.error(f"Erro ao processar inventário: {e}")
        return {
            'error': str(e),
            'itens': [],
            'total': 0,
            'abas_disponiveis': []
        }


def get_abas_disponiveis():
    """Retorna apenas a lista de abas da planilha de inventário."""
    if not INVENTARIO_URL:
        return {'error': 'URL do inventário não configurada', 'abas': []}
    try:
        response = requests.get(INVENTARIO_URL, timeout=30)
        response.raise_for_status()
        wb = load_workbook(BytesIO(response.content), data_only=True)
        return {'abas': wb.sheetnames}
    except Exception as e:
        logger.error(f"Erro ao buscar abas: {e}")
        return {'error': str(e), 'abas': []}
